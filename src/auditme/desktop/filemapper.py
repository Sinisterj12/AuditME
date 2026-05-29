#!/usr/bin/env python3
"""
FileMapper — Lightweight desktop viewer for color-coded project skeletons.

Reads .xlsx workbooks containing "Control Map" + "Legend" tabs and renders
them as a beautiful, zoomable, filterable, editable IDE-style tree-view app
specifically optimized for AI Pair Programming, Vibe Coding, and ADHD scanability.

Keyboard shortcuts:
  Ctrl+O       Open file
  Ctrl+S       Save changes
  Ctrl+F       Focus search
  Ctrl+= / -   Zoom in / out
  Ctrl+0       Reset zoom
  1 to 9       Quick-assign status to selected item
  (Letters: c, n, s, a, f, h, t, r, l also work)
"""

import sys
import os
import json
import math
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, font as tkfont, colorchooser
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    print("ERROR: openpyxl is required.  pip install openpyxl")
    sys.exit(1)


# ═══════════════════════════════════════════════════════════════════════
# Config persistence
# ═══════════════════════════════════════════════════════════════════════
CONFIG_PATH = Path(__file__).parent / "filemapper_config.json"


# ═══════════════════════════════════════════════════════════════════════
# STATUS ICON SYSTEM — unambiguous, workflow-accurate icons
# ═══════════════════════════════════════════════════════════════════════
#
# Design principle: Each icon should INSTANTLY communicate meaning
# even without reading the label. No more "filled vs empty circle" confusion.
#
# Workflow order (left = do first, right = done):
#   🔨 BUILD → ⏭ QUEUE → 🚀 SHIP → ✅ ACTIVE → 🔒 FREEZE
#
STATUS_ICONS = {
    # ── Active work states ──────────────────────────────────────────────
    "CONSTRUCTION":  "🔨",   # Hammer = being hammered out RIGHT NOW
    "NEXT":          "⏭",   # Skip-forward = queued up, play it next
    "PLACEHOLDER":   "📋",   # Clipboard = placeholder, fill this in later

    # ── Live/stable states ──────────────────────────────────────────────
    "ACTIVE_V1":     "⚡",   # Lightning bolt = LIVE and running in production
    "SHIP_READY":    "✅",   # Checkmark = verified, done, safe
    "FROZEN":        "🔒",   # Padlock = locked, do not change

    # ── Human/gate states ───────────────────────────────────────────────
    "HUMAN_GATE":    "🚧",   # Construction barrier = STOP, needs human decision

    # ── Archive states ──────────────────────────────────────────────────
    "ROADMAP":       "🗺",   # Map = future destination, not here yet
    "LEGACY":        "📦",   # Box = packed away, archive only
}

# Text labels for the workflow guide (shown in legend)
STATUS_WORKFLOW = {
    "CONSTRUCTION":  ("🔨 BUILD",    "Working on it NOW",           "BUILD → Ship when done"),
    "NEXT":          ("⏭ QUEUE",     "Approved, up next",           "Start after current build finishes"),
    "PLACEHOLDER":   ("📋 SKELETON", "Placeholder exists",          "Fill in the implementation"),
    "ACTIVE_V1":     ("⚡ LIVE",     "Running in production V1",    "Use it — improve carefully"),
    "SHIP_READY":    ("✅ DONE",     "Verified & safe",             "Do NOT modify"),
    "FROZEN":        ("🔒 LOCKED",   "Locked — no changes allowed", "Requires decision log to unlock"),
    "HUMAN_GATE":    ("🚧 GATE",     "Needs James to approve",      "STOP — get approval first"),
    "ROADMAP":       ("🗺 FUTURE",   "Future — off limits now",     "Do NOT build yet"),
    "LEGACY":        ("📦 ARCHIVE",  "Old code — archive only",     "Mine for reference only"),
}

STATUS_ORDER = [
    "CONSTRUCTION", "NEXT", "PLACEHOLDER",
    "ACTIVE_V1", "SHIP_READY", "FROZEN",
    "HUMAN_GATE", "ROADMAP", "LEGACY"
]

STATUS_META = {
    "CONSTRUCTION":  {"label": "Orange",       "meaning": "Being built now",      "action": "Current work area."},
    "NEXT":          {"label": "Yellow",       "meaning": "Queued next",          "action": "Start after current build."},
    "PLACEHOLDER":   {"label": "Gold",         "meaning": "Placeholder",          "action": "Create now, fill later."},
    "ACTIVE_V1":     {"label": "Blue",         "meaning": "Live V1",              "action": "Use and improve carefully."},
    "SHIP_READY":    {"label": "Green",        "meaning": "Verified and done",    "action": "Safe — do NOT modify."},
    "FROZEN":        {"label": "Light Blue",   "meaning": "Locked",               "action": "Do not change without log."},
    "HUMAN_GATE":    {"label": "Purple",       "meaning": "Needs human approval", "action": "STOP — get approval first."},
    "ROADMAP":       {"label": "Red",          "meaning": "Future / off-limits",  "action": "Do not build until promoted."},
    "LEGACY":        {"label": "Gray",         "meaning": "Archive only",         "action": "Mine later; do not operate."},
}

# Disk existence indicators — clearly NOT circles to avoid confusion
DISK_ICONS = {
    "exists": "💾",   # Floppy = file IS on disk
    "missing": "✗",   # X = file NOT on disk, needs stub
}


# ═══════════════════════════════════════════════════════════════════════
# COLOR THEME PRESETS
# ═══════════════════════════════════════════════════════════════════════
COLOR_THEMES = {
    "🎨 Vivid (Default)": {
        "CONSTRUCTION":  "#FF8C00",   # Deep orange
        "NEXT":          "#FFD700",   # Gold yellow
        "PLACEHOLDER":   "#FFC107",   # Amber — clearly different from orange
        "ACTIVE_V1":     "#2196F3",   # Material blue
        "SHIP_READY":    "#4CAF50",   # Material green
        "FROZEN":        "#80CBC4",   # Teal-mint — very different from blue
        "HUMAN_GATE":    "#9C27B0",   # Deep purple
        "ROADMAP":       "#F44336",   # Material red
        "LEGACY":        "#9E9E9E",   # Medium gray
    },
    "👁 High Contrast": {
        "CONSTRUCTION":  "#E65100",   # Very dark orange
        "NEXT":          "#F9A825",   # Dark amber
        "PLACEHOLDER":   "#FFD54F",   # Light amber
        "ACTIVE_V1":     "#0D47A1",   # Very dark blue
        "SHIP_READY":    "#1B5E20",   # Very dark green
        "FROZEN":        "#00ACC1",   # Vivid cyan
        "HUMAN_GATE":    "#6A1B9A",   # Very dark purple
        "ROADMAP":       "#B71C1C",   # Very dark red
        "LEGACY":        "#616161",   # Dark gray
    },
    "🟢 Color-Blind Safe (Deuteranopia)": {
        # Avoids red-green confusion — uses blue/orange/violet spectrum
        "CONSTRUCTION":  "#E69F00",   # Orange (Brewer safe)
        "NEXT":          "#F0E442",   # Yellow (Brewer safe)
        "PLACEHOLDER":   "#D55E00",   # Vermillion (Brewer safe)
        "ACTIVE_V1":     "#0072B2",   # Blue (Brewer safe)
        "SHIP_READY":    "#009E73",   # Bluish-green (Brewer safe)
        "FROZEN":        "#56B4E9",   # Sky blue (Brewer safe)
        "HUMAN_GATE":    "#CC79A7",   # Mauve/pink (Brewer safe)
        "ROADMAP":       "#999999",   # Gray (avoids red)
        "LEGACY":        "#CCCCCC",   # Light gray
    },
    "🌸 Pastel (Easy on eyes)": {
        "CONSTRUCTION":  "#FFCCBC",   # Peach
        "NEXT":          "#FFF9C4",   # Pale yellow
        "PLACEHOLDER":   "#FFE082",   # Pale amber
        "ACTIVE_V1":     "#BBDEFB",   # Pale blue
        "SHIP_READY":    "#C8E6C9",   # Pale green
        "FROZEN":        "#E0F7FA",   # Pale cyan
        "HUMAN_GATE":    "#E1BEE7",   # Pale purple
        "ROADMAP":       "#FFCDD2",   # Pale red
        "LEGACY":        "#F5F5F5",   # Off white
    },
    "🌟 Neon Dark": {
        "CONSTRUCTION":  "#FF6B00",   # Neon orange
        "NEXT":          "#FFE600",   # Neon yellow
        "PLACEHOLDER":   "#FF9F00",   # Neon amber
        "ACTIVE_V1":     "#00B0FF",   # Neon blue
        "SHIP_READY":    "#00E676",   # Neon green
        "FROZEN":        "#00E5FF",   # Neon cyan
        "HUMAN_GATE":    "#D500F9",   # Neon purple
        "ROADMAP":       "#FF1744",   # Neon red
        "LEGACY":        "#757575",   # Muted gray
    },
}

THEME_NAMES = list(COLOR_THEMES.keys())


def _load_config() -> dict:
    """Load persisted config (selected theme + any custom colors + project roots)."""
    if CONFIG_PATH.exists():
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
                if "project_roots" not in data:
                    data["project_roots"] = {}
                if "stub_threshold_bytes" not in data:
                    data["stub_threshold_bytes"] = 400
                return data
        except Exception:
            pass
    return {"theme": THEME_NAMES[0], "custom_colors": {}, "project_roots": {}, "stub_threshold_bytes": 400}


def _save_config(data: dict):
    try:
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
    except Exception:
        pass


# ═══════════════════════════════════════════════════════════════════════
# UI Dark Theme
# ═══════════════════════════════════════════════════════════════════════
DARK = {
    "bg":           "#07101E",   # Near-black deep cosmic navy
    "surface":      "#0F1C2E",   # Rich dark surface
    "card":         "#162035",   # Dark blue-slate card
    "text":         "#E8F0FE",   # Crisp cool white
    "text_dim":     "#6B839A",   # Muted blue-gray
    "accent":       "#29B6F6",   # Sky blue
    "accent_hover": "#81D4FA",   # Lighter sky
    "border":       "#1E3050",   # Subtle border
    "header_bg":    "#0A1628",   # Deep header
    "header_fg":    "#E8F0FE",
    "danger":       "#EF5350",
    "success":      "#26A69A",
    "warn":         "#FFA726",
    "tree_selected":"#1A3A6C",
}


# ═══════════════════════════════════════════════════════════════════════
# Scrollable Frame
# ═══════════════════════════════════════════════════════════════════════
class ScrollableFrame(tk.Frame):
    def __init__(self, container, bg=None, *args, **kwargs):
        bg = bg or DARK["bg"]
        super().__init__(container, *args, bg=bg, **kwargs)
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        self.canvas = tk.Canvas(self, bg=bg, borderwidth=0, highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview,
                                       style="Dk.Vertical.TScrollbar")
        self.scrollable_frame = tk.Frame(self.canvas, bg=bg)
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas_frame = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.bind("<Configure>", lambda e: self.canvas.itemconfig(self.canvas_frame, width=e.width))

        def _mw(event): self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        self.canvas.bind("<Enter>",  lambda e: self.canvas.bind_all("<MouseWheel>", _mw))
        self.canvas.bind("<Leave>",  lambda e: self.canvas.unbind_all("<MouseWheel>"))
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")


# ═══════════════════════════════════════════════════════════════════════
# Main Application
# ═══════════════════════════════════════════════════════════════════════
class FileMapperApp:

    def __init__(self, root: tk.Tk, xlsx_path: str | None = None):
        self.root = root
        self.root.title("AuditME Desktop — Visual Cockpit & Agent Workday Dashboard")
        self.root.geometry("1760x1000")
        self.root.minsize(1280, 700)
        self.root.configure(bg=DARK["bg"])

        # ── Load config ──
        self._config = _load_config()
        self._theme_name: str = self._config.get("theme", THEME_NAMES[0])
        self._custom_colors: dict = self._config.get("custom_colors", {})
        self.status_colors: dict = self._effective_colors()

        # ── State ──
        if not xlsx_path:
            # Auto-discovery
            root_candidate = Path(".").resolve()
            while root_candidate != root_candidate.parent:
                if (root_candidate / "90_AUDITME").is_dir() or (root_candidate / ".git").is_dir():
                    xlsx_path = str(root_candidate)
                    break
                root_candidate = root_candidate.parent
        self.xlsx_path: str | None = xlsx_path
        self.repository_mode: bool = False
        self.workbook = None
        self.rows: list[dict] = []
        self.zoom_level: int = 0
        self._base_size: int = 11
        self.active_filter: str | None = None
        self.dirty: bool = False
        self.current_editor_idx: int | None = None
        self._bg_pulse = 0
        self.linked_project_root: str | None = None
        self.git_changed_files: dict[str, str] = {}
        self.current_git_branch: str | None = None
        self.branch_diff_files: set[str] = set()
        self.git_head_sizes: dict[str, int] = {}

        # ── Workday State ──
        workday_cfg = self._config.get("workday", {})
        self.workday_active = workday_cfg.get("active", False)
        self.workday_start_time = workday_cfg.get("start_time", None)
        self.workday_duration_mins = workday_cfg.get("duration_mins", 60)
        self.workday_completed_tasks = workday_cfg.get("completed_tasks", [])
        self.workday_task_start_times = workday_cfg.get("task_start_times", {})
        
        # Collapsible Preference States
        self.git_sync_collapsed = True
        self.shortcuts_collapsed = True
        self.unlocked_frozen_paths = set()
        self.collapsed_sections = {
            "workday": False,
            "copier": False,
            "recommend": False,
            "recent": False,
            "todo": False,
            "horizon": False
        }

        # ── Fonts ──
        self._family = "Segoe UI"
        self._mono   = "Cascadia Code"
        self._build_fonts()

        # ── Theme / style ──
        self.style = ttk.Style()
        self._apply_theme()

        # ── Build layout ──
        self._build_chrome()
        self._build_content()
        self._build_statusbar()

        # ── Key bindings ──
        self.root.bind("<Control-plus>",  lambda e: self._zoom(1))
        self.root.bind("<Control-equal>", lambda e: self._zoom(1))
        self.root.bind("<Control-minus>", lambda e: self._zoom(-1))
        self.root.bind("<Control-0>",     lambda e: self._zoom_reset())
        self.root.bind("<Control-o>",     lambda e: self._open_file())
        self.root.bind("<Control-Shift-O>", lambda e: self._open_directory())
        self.root.bind("<Control-s>",     lambda e: self._save())
        self.root.bind("<Control-f>",     lambda e: self._focus_search())
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

        # ── Context menu ──
        self.menu = tk.Menu(self.root, tearoff=0, bg=DARK["card"], fg=DARK["text"],
                            activebackground=DARK["accent"], activeforeground="#0A1628",
                            relief="flat", font=self.f)

        # ── ADHD rotating tips ──
        self.tips = [
            "🎯 Focus Mode: Hide all non-active files instantly — click 🎯 Focus Mode toggle in the toolbar.",
            "⚡ Quickfire: Press 1–9 or C/N/S/A/F/H/T/R/L to set status on selected row in one keystroke.",
            "🎨 Change color themes: Click '🎨 Theme' in the toolbar to cycle between 5 preset palettes.",
            "🔨 CONSTRUCTION = you're building it RIGHT NOW.  ⏭ NEXT = it's queued but not started.",
            "⚡ ACTIVE_V1 = it's LIVE and working — not empty, it means it's running in production!",
            "📋 STUB = placeholder file exists on disk but has no logic yet — you need to fill it in.",
            "✅ SHIP_READY = verified, done, do not touch.  🔒 FROZEN = locked, requires a decision log.",
            "🚧 HUMAN_GATE = STOP — this requires a human approval before work can continue.",
            "📋 Select any file → click 'Copy AI Prompt' to get an optimized implementation prompt.",
            "🌳 Use ⊞ All / ⊟ All buttons to instantly expand or collapse the entire tree.",
        ]
        self.current_tip_idx = 0
        self._rotate_tips()
        self._animate_bg()

        # ── Auto-load ──
        if self.xlsx_path:
            self._load(self.xlsx_path)
        else:
            self.root.after(200, self._open_file)

        # Periodic Git Sync (every 30 seconds)
        self.root.after(30000, self._periodic_git_refresh)
        
        # Start Workday Shift real-time update loop
        self._update_workday_timer_ui()

    # ─────────────────────────────────────────────────────────────────
    # Color helpers
    # ─────────────────────────────────────────────────────────────────
    def _effective_colors(self) -> dict:
        """Merge theme preset + any custom overrides."""
        base = COLOR_THEMES.get(self._theme_name, COLOR_THEMES[THEME_NAMES[0]]).copy()
        base.update(self._custom_colors)
        return base

    def _apply_color_theme(self, theme_name: str):
        self._theme_name = theme_name
        self._custom_colors = {}           # reset custom when switching preset
        self.status_colors = self._effective_colors()
        self._persist_config()
        self._reconfigure_tree_tags()
        self._rebuild_filter_pills()
        self._rebuild_legend_rows()
        self._repopulate()
        self.sb_left.config(text=f"🎨 Theme applied: {theme_name}")

    def _persist_config(self):
        self._config["theme"]         = self._theme_name
        self._config["custom_colors"] = self._custom_colors
        self._config["workday"] = {
            "active": self.workday_active,
            "start_time": self.workday_start_time,
            "duration_mins": self.workday_duration_mins,
            "completed_tasks": self.workday_completed_tasks,
            "task_start_times": self.workday_task_start_times,
        }
        _save_config(self._config)

    # ─────────────────────────────────────────────────────────────────
    # Fonts
    # ─────────────────────────────────────────────────────────────────
    def _build_fonts(self):
        s = max(8, self._base_size + self.zoom_level)
        self.f    = tkfont.Font(family=self._family, size=s)
        self.fb   = tkfont.Font(family=self._family, size=s, weight="bold")
        self.fh   = tkfont.Font(family=self._family, size=s + 2, weight="bold")
        self.fs   = tkfont.Font(family=self._family, size=max(7, s - 2))
        self.fm   = tkfont.Font(family=self._mono,   size=s)
        self.fmb  = tkfont.Font(family=self._mono,   size=s, weight="bold")
        self.ftree = tkfont.Font(family=self._mono,  size=max(9, s))

    def _refresh_fonts(self):
        s = max(8, self._base_size + self.zoom_level)
        for fnt, delta, w in [(self.f,  0, "normal"), (self.fb, 0, "bold"),
                               (self.fh, 2, "bold"),   (self.fs, -2, "normal"),
                               (self.fm, 0, "normal"), (self.fmb, 0, "bold"),
                               (self.ftree, 0, "normal")]:
            fnt.configure(size=max(7, s + delta), weight=w)
        self.style.configure("FM.Treeview", rowheight=max(28, s + 18))
        self.zoom_lbl.config(text=f"{self.zoom_level:+d}")

    # ─────────────────────────────────────────────────────────────────
    # Animated background pulse
    # ─────────────────────────────────────────────────────────────────
    def _animate_bg(self):
        t = self._bg_pulse
        f = (math.sin(t * 0.03) + 1) / 2
        r = int(0x0F + (0x13 - 0x0F) * f)
        g = int(0x1C + (0x22 - 0x1C) * f)
        b = int(0x2E + (0x38 - 0x2E) * f)
        col = f"#{r:02x}{g:02x}{b:02x}"
        if hasattr(self, "_chrome_bar"):
            try:
                self._chrome_bar.configure(bg=col)
                for w in self._chrome_bar.winfo_children():
                    if isinstance(w, tk.Label):
                        w.configure(bg=col)
                    elif isinstance(w, tk.Frame):
                        w.configure(bg=col)
                        for ww in w.winfo_children():
                            if isinstance(ww, tk.Label):
                                ww.configure(bg=col)
            except Exception:
                pass
        self._bg_pulse += 1
        self.root.after(80, self._animate_bg)

    # ─────────────────────────────────────────────────────────────────
    # ttk Style
    # ─────────────────────────────────────────────────────────────────
    def _apply_theme(self):
        self.style.theme_use("clam")
        D = DARK

        self.style.configure("FM.Treeview",
            background=D["surface"], foreground=D["text"],
            fieldbackground=D["surface"], borderwidth=0, rowheight=30,
            font=(self._mono, 10))
        self.style.configure("FM.Treeview.Heading",
            background=D["header_bg"], foreground=D["header_fg"],
            font=(self._family, 9, "bold"), borderwidth=0, relief="flat")
        self.style.map("FM.Treeview",
            background=[("selected", D["tree_selected"])],
            foreground=[("selected", "#C8E6FF")])
        self.style.map("FM.Treeview.Heading",
            background=[("active", "#112040")])

        self.style.configure("TCombobox",
            fieldbackground=D["surface"], background=D["card"], foreground=D["text"],
            bordercolor=D["border"], arrowcolor=D["text"])
        self.style.map("TCombobox",
            fieldbackground=[("readonly", D["surface"])],
            foreground=[("readonly", D["text"])])

        for name, bg, fg in [
            ("Dk.TButton", D["card"],   D["text"]),
            ("Ac.TButton", D["accent"], "#07101E"),
        ]:
            self.style.configure(name, background=bg, foreground=fg,
                borderwidth=1, relief="flat", padding=(12, 6), font=self.f)
        self.style.map("Dk.TButton",
            background=[("active", D["accent"]), ("pressed", D["accent_hover"])],
            foreground=[("active", "#07101E")])
        self.style.map("Ac.TButton",
            background=[("active", D["accent_hover"])])

        for n, bg in [("Bg.TFrame", D["bg"]), ("Card.TFrame", D["card"]), ("Surf.TFrame", D["surface"])]:
            self.style.configure(n, background=bg)

        for n in ["Dk.Vertical.TScrollbar", "Dk.Horizontal.TScrollbar"]:
            self.style.configure(n, background=D["card"], troughcolor=D["surface"],
                borderwidth=0, arrowcolor=D["text_dim"])

    # ─────────────────────────────────────────────────────────────────
    # Tree tag color config
    # ─────────────────────────────────────────────────────────────────
    def _reconfigure_tree_tags(self):
        for status in STATUS_META:
            bg = self.status_colors.get(status, "#333")
            # For neon themes use dark text; for pastels use near-black
            r, g, b = int(bg[1:3], 16), int(bg[3:5], 16), int(bg[5:7], 16)
            luminance = 0.2126 * r + 0.7152 * g + 0.0722 * b
            fg = "#0A0A0A" if luminance > 80 else "#1A1A1A"
            self.tree.tag_configure(status, background=bg, foreground=fg)
        self.tree.tag_configure("virtual_folder",
            background=DARK["header_bg"], foreground=DARK["text_dim"])
        self.tree.tag_configure("workspace_root",
            background=DARK["card"], foreground=DARK["accent"])
        self.tree.tag_configure("git_modified", foreground="#FFD700")
        self.tree.tag_configure("git_untracked", foreground="#00E676")
        self.tree.tag_configure("not_in_main", foreground="#FF9F00")

    # ─────────────────────────────────────────────────────────────────
    # Chrome (top bar)
    # ─────────────────────────────────────────────────────────────────
    def _build_chrome(self):
        bar = tk.Frame(self.root, bg=DARK["surface"], height=58)
        bar.pack(fill=tk.X)
        bar.pack_propagate(False)
        self._chrome_bar = bar

        # Left accent stripe
        tk.Frame(bar, bg=DARK["accent"], width=4).pack(side=tk.LEFT, fill=tk.Y)

        # Branding
        tk.Label(bar, text="📁", font=("Segoe UI Emoji", 20),
                 bg=DARK["surface"], fg=DARK["accent"]).pack(side=tk.LEFT, padx=(12, 4))
        tk.Label(bar, text="AuditME Desktop",
                 font=(self._family, 16, "bold"),
                 bg=DARK["surface"], fg=DARK["text"]).pack(side=tk.LEFT)
        self.subtitle_lbl = tk.Label(bar, text="  No file loaded",
                 font=(self._family, 9), bg=DARK["surface"], fg=DARK["text_dim"])
        self.subtitle_lbl.pack(side=tk.LEFT, padx=(8, 0))

        # Project link button
        self.link_lbl = tk.Button(bar, text="🔗 Link Project Root",
                                  font=(self._family, 9, "bold"),
                                  bg=DARK["surface"], fg=DARK["accent"],
                                  activebackground=DARK["surface"], activeforeground=DARK["accent_hover"],
                                  relief="flat", bd=0, cursor="hand2",
                                  command=self._on_link_project_click)
        self.link_lbl.pack(side=tk.LEFT, padx=(15, 0))

        # Right actions
        rf = tk.Frame(bar, bg=DARK["surface"])
        rf.pack(side=tk.RIGHT, padx=12)

        ttk.Button(rf, text="📂 Open Excel",    style="Dk.TButton", command=self._open_file).pack(side=tk.LEFT, padx=2)
        ttk.Button(rf, text="🌿 Open Repo",     style="Dk.TButton", command=self._open_directory).pack(side=tk.LEFT, padx=2)
        ttk.Button(rf, text="💾 Save",           style="Dk.TButton", command=self._save).pack(side=tk.LEFT, padx=2)
        ttk.Button(rf, text="🤖 LLM Export",     style="Ac.TButton", command=self._export_llm_roadmap).pack(side=tk.LEFT, padx=2)

        # Theme switcher button
        self.theme_lbl_var = tk.StringVar(value=f"🎨 {self._theme_name.split(' ', 1)[0]}")
        theme_btn = tk.Button(rf, textvariable=self.theme_lbl_var,
                              font=(self._family, 9, "bold"),
                              bg=DARK["card"], fg=DARK["accent"],
                              relief="flat", bd=1, padx=10, pady=5, cursor="hand2",
                              activebackground=DARK["accent"], activeforeground="#07101E",
                              highlightthickness=1, highlightbackground=DARK["accent"],
                              command=self._show_theme_picker)
        theme_btn.pack(side=tk.LEFT, padx=4)

        # Zoom
        zf = tk.Frame(bar, bg=DARK["surface"])
        zf.pack(side=tk.RIGHT, padx=8)
        tk.Label(zf, text="Zoom", font=(self._family, 8),
                 bg=DARK["surface"], fg=DARK["text_dim"]).pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(zf, text="−", style="Dk.TButton", width=3, command=lambda: self._zoom(-1)).pack(side=tk.LEFT)
        self.zoom_lbl = tk.Label(zf, text="+0", width=4, font=(self._family, 9),
                                 bg=DARK["surface"], fg=DARK["text_dim"])
        self.zoom_lbl.pack(side=tk.LEFT, padx=3)
        ttk.Button(zf, text="+", style="Dk.TButton", width=3, command=lambda: self._zoom(1)).pack(side=tk.LEFT)

    # ─────────────────────────────────────────────────────────────────
    # Theme Picker Dialog
    # ─────────────────────────────────────────────────────────────────
    def _show_theme_picker(self):
        dlg = tk.Toplevel(self.root)
        dlg.title("🎨 Choose Color Theme")
        dlg.geometry("540x560")
        dlg.configure(bg=DARK["bg"])
        dlg.transient(self.root)
        dlg.grab_set()

        tk.Label(dlg, text="🎨 Color Theme Presets",
                 font=(self._family, 14, "bold"), bg=DARK["bg"], fg=DARK["accent"]
                 ).pack(anchor=tk.W, padx=20, pady=(16, 4))
        tk.Label(dlg, text="Choose a preset palette — colors are tuned for clarity, contrast, and accessibility.",
                 font=(self._family, 9), bg=DARK["bg"], fg=DARK["text_dim"]
                 ).pack(anchor=tk.W, padx=20, pady=(0, 12))

        for theme_name, colors in COLOR_THEMES.items():
            is_active = (theme_name == self._theme_name)
            row_bg = DARK["card"] if is_active else DARK["surface"]
            border = DARK["accent"] if is_active else DARK["border"]

            row = tk.Frame(dlg, bg=row_bg, highlightthickness=2,
                           highlightbackground=border, cursor="hand2")
            row.pack(fill=tk.X, padx=16, pady=4)

            # Theme label
            lbl = tk.Label(row, text=theme_name + (" ✓ Active" if is_active else ""),
                           font=(self._family, 10, "bold" if is_active else "normal"),
                           bg=row_bg, fg=DARK["accent"] if is_active else DARK["text"],
                           anchor=tk.W)
            lbl.pack(side=tk.LEFT, padx=10, pady=8)

            # Color swatches preview
            swatches = tk.Frame(row, bg=row_bg)
            swatches.pack(side=tk.RIGHT, padx=10, pady=6)
            for status in STATUS_ORDER[:6]:
                sw_bg = colors.get(status, "#888")
                sw = tk.Canvas(swatches, width=22, height=22, bg=row_bg, highlightthickness=1,
                               highlightbackground=DARK["border"])
                sw.pack(side=tk.LEFT, padx=1)
                sw.create_rectangle(2, 2, 20, 20, fill=sw_bg, outline="")
                # Draw the icon on the swatch
                icon = STATUS_ICONS.get(status, "?")
                sw.create_text(11, 11, text=icon, font=("Segoe UI Emoji", 9))

            def _apply(n=theme_name, d=dlg):
                self._apply_color_theme(n)
                self.theme_lbl_var.set(f"🎨 {n.split(' ', 1)[0]}")
                d.destroy()

            row.bind("<Button-1>", lambda e, n=theme_name: _apply(n))
            lbl.bind("<Button-1>", lambda e, n=theme_name: _apply(n))

        # Custom color picker option
        tk.Frame(dlg, bg=DARK["border"], height=1).pack(fill=tk.X, padx=16, pady=8)
        tk.Label(dlg, text="Or customize individual status colors:",
                 font=(self._family, 9), bg=DARK["bg"], fg=DARK["text_dim"]).pack(anchor=tk.W, padx=20)
        tk.Label(dlg, text="→ Switch to the Stats & Legend tab and click the color swatch next to any status.",
                 font=(self._family, 8), bg=DARK["bg"], fg=DARK["accent"]).pack(anchor=tk.W, padx=28, pady=(0, 10))

        tk.Button(dlg, text="Close", font=(self._family, 10),
                  bg=DARK["card"], fg=DARK["text"], relief="flat", bd=1, padx=20, pady=8,
                  cursor="hand2", command=dlg.destroy).pack(side=tk.BOTTOM, pady=12)

    # ─────────────────────────────────────────────────────────────────
    # Content area
    # ─────────────────────────────────────────────────────────────────
    def _build_content(self):
        wrap = tk.Frame(self.root, bg=DARK["bg"])
        wrap.pack(fill=tk.BOTH, expand=True)

        # Left — tree panel
        left = tk.Frame(wrap, bg=DARK["bg"])
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(8, 4), pady=4)

        self._build_filters(left)
        self._build_workflow_guide(left)
        self._build_tree(left)

        # Tip banner
        self.tip_frame = tk.Frame(left, bg=DARK["surface"], height=28,
                                  highlightthickness=1, highlightbackground=DARK["border"])
        self.tip_frame.pack(fill=tk.X, pady=(4, 0))
        self.tip_frame.pack_propagate(False)
        tk.Label(self.tip_frame, text="●", font=("Segoe UI", 8),
                 bg=DARK["surface"], fg=DARK["accent"]).pack(side=tk.LEFT, padx=(8, 4))
        self.tip_lbl = tk.Label(self.tip_frame, text="💡 Loading...",
                                font=(self._family, 8), bg=DARK["surface"],
                                fg=DARK["text_dim"], anchor=tk.W)
        self.tip_lbl.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 8))

        # Right — sidebar
        right = tk.Frame(wrap, bg=DARK["bg"], width=440)
        right.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 8), pady=4)
        right.pack_propagate(False)
        self._build_right_sidebar(right)

    # ─────────────────────────────────────────────────────────────────
    # Workflow guide strip (always visible above the tree)
    # ─────────────────────────────────────────────────────────────────
    def _build_workflow_guide(self, parent):
        """A single-line strip showing the workflow in order — always visible."""
        strip = tk.Frame(parent, bg=DARK["header_bg"], height=32,
                         highlightthickness=1, highlightbackground=DARK["border"])
        strip.pack(fill=tk.X, pady=(2, 0))
        strip.pack_propagate(False)

        tk.Label(strip, text="Work Order →", font=(self._family, 7, "bold"),
                 bg=DARK["header_bg"], fg=DARK["text_dim"]).pack(side=tk.LEFT, padx=(8, 6))

        # Show the active-work statuses in order
        priority_statuses = ["CONSTRUCTION", "NEXT", "PLACEHOLDER", "ACTIVE_V1", "SHIP_READY", "FROZEN"]
        for i, status in enumerate(priority_statuses):
            icon  = STATUS_ICONS[status]
            short, _, tip = STATUS_WORKFLOW[status]
            bg    = self.status_colors.get(status, "#888")

            r, g, b = int(bg[1:3], 16), int(bg[3:5], 16), int(bg[5:7], 16)
            lum = 0.2126*r + 0.7152*g + 0.0722*b
            fg = "#0A0A0A" if lum > 80 else "#1A1A1A"

            chip = tk.Frame(strip, bg=bg, highlightthickness=0)
            chip.pack(side=tk.LEFT, padx=2, pady=4)

            tk.Label(chip, text=f" {icon} {short} ",
                     font=(self._family, 7, "bold"), bg=bg, fg=fg
                     ).pack(side=tk.LEFT, padx=2)

            if i < len(priority_statuses) - 1:
                tk.Label(strip, text="→", font=(self._family, 8),
                         bg=DARK["header_bg"], fg=DARK["text_dim"]).pack(side=tk.LEFT)

        # Right: help label
        tk.Label(strip, text="(What to touch next →)",
                 font=(self._family, 7), bg=DARK["header_bg"], fg=DARK["text_dim"]
                 ).pack(side=tk.RIGHT, padx=8)

    # ─────────────────────────────────────────────────────────────────
    # Filter pills
    # ─────────────────────────────────────────────────────────────────
    def _build_filters(self, parent):
        bar = tk.Frame(parent, bg=DARK["bg"])
        bar.pack(fill=tk.X, pady=(0, 2))

        tk.Label(bar, text="🔍", font=("Segoe UI Emoji", 11),
                 bg=DARK["bg"], fg=DARK["text_dim"]).pack(side=tk.LEFT, padx=(0, 4))
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *_: self._repopulate())
        self.search_entry = tk.Entry(bar, textvariable=self.search_var,
            font=self.f, bg=DARK["card"], fg=DARK["text"],
            insertbackground=DARK["text"], relief="flat", bd=0,
            highlightthickness=1, highlightcolor=DARK["accent"],
            highlightbackground=DARK["border"])
        self.search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=6, padx=(0, 8))

        tk.Button(bar, text="➕ Add Node", font=(self._family, 8, "bold"),
                  bg=DARK["accent"], fg="#07101E", relief="flat", bd=0, padx=10, pady=5,
                  activebackground=DARK["accent_hover"], cursor="hand2",
                  command=self._on_add_node_click).pack(side=tk.LEFT, padx=(0, 6))

        self.focus_var = tk.BooleanVar(value=False)
        tk.Checkbutton(bar, text="🎯 Focus", variable=self.focus_var,
                       font=(self._family, 8, "bold"), bg=DARK["bg"], fg=DARK["accent"],
                       activebackground=DARK["bg"], activeforeground=DARK["accent_hover"],
                       selectcolor=DARK["surface"], relief="flat", cursor="hand2",
                       command=self._repopulate).pack(side=tk.LEFT, padx=(0, 4))

        tk.Button(bar, text="⊞", font=(self._family, 9), bg=DARK["card"], fg=DARK["accent"],
                  relief="flat", bd=0, padx=6, pady=4, cursor="hand2",
                  activebackground=DARK["accent"], activeforeground="#07101E",
                  command=self._expand_all_nodes).pack(side=tk.LEFT, padx=1)
        tk.Button(bar, text="⊟", font=(self._family, 9), bg=DARK["card"], fg=DARK["text_dim"],
                  relief="flat", bd=0, padx=6, pady=4, cursor="hand2",
                  activebackground=DARK["accent"], activeforeground="#07101E",
                  command=self._collapse_all_nodes).pack(side=tk.LEFT, padx=1)

        # Pills strip
        self._pill_bar = tk.Frame(parent, bg=DARK["bg"])
        self._pill_bar.pack(fill=tk.X, pady=(2, 4))
        self.pills: dict[str, tk.Button] = {}
        self._rebuild_filter_pills()

    def _rebuild_filter_pills(self):
        for w in self._pill_bar.winfo_children():
            w.destroy()
        self.pills = {}

        # ALL pill
        all_btn = tk.Button(self._pill_bar, text="📁 All",
            font=(self._family, 7, "bold"), bg=DARK["accent"], fg="#07101E",
            relief="flat", bd=0, padx=8, pady=3, cursor="hand2",
            activebackground=DARK["accent_hover"],
            command=lambda: self._set_filter("ALL"))
        all_btn.pack(side=tk.LEFT, padx=1)
        self.pills["ALL"] = all_btn

        for status in STATUS_ORDER:
            bg   = self.status_colors.get(status, "#888")
            icon = STATUS_ICONS.get(status, "?")
            short = STATUS_WORKFLOW[status][0].split(" ", 1)[1]  # e.g. "BUILD"
            if len(short) > 8:
                short = short[:7] + "…"
            btn = tk.Button(self._pill_bar, text=f"{icon} {short}",
                font=(self._family, 7, "bold"), bg=bg, fg="#0A0A0A",
                relief="flat", bd=0, padx=6, pady=3, cursor="hand2",
                activebackground=DARK["accent_hover"],
                command=lambda s=status: self._set_filter(s))
            btn.pack(side=tk.LEFT, padx=1)
            self.pills[status] = btn

    # ─────────────────────────────────────────────────────────────────
    # Treeview
    # ─────────────────────────────────────────────────────────────────
    def _build_tree(self, parent):
        hf = tk.Frame(parent, bg=DARK["header_bg"], height=28)
        hf.pack(fill=tk.X)
        hf.pack_propagate(False)
        tk.Label(hf, text="EXPLORER", font=(self._family, 8, "bold"),
                 bg=DARK["header_bg"], fg=DARK["text_dim"], anchor=tk.W
                 ).pack(side=tk.LEFT, padx=10, fill=tk.Y)
        self.tree_count_lbl = tk.Label(hf, text="",
                 font=(self._family, 8), bg=DARK["header_bg"], fg=DARK["accent"], anchor=tk.E)
        self.tree_count_lbl.pack(side=tk.RIGHT, padx=10, fill=tk.Y)

        tf = tk.Frame(parent, bg=DARK["surface"],
                      highlightthickness=1, highlightbackground=DARK["border"])
        tf.pack(fill=tk.BOTH, expand=True)

        self.tree = ttk.Treeview(tf, columns=["Status"],
            show="tree headings", style="FM.Treeview", selectmode="extended")
        self.tree.heading("#0",     text="  File / Folder Structure", anchor=tk.W)
        self.tree.column("#0",      width=520, minwidth=280, stretch=True)
        self.tree.heading("Status", text="Status  (icon = what it is)", anchor=tk.W)
        self.tree.column("Status",  width=180, minwidth=100, stretch=False)

        vsb = ttk.Scrollbar(tf, orient=tk.VERTICAL,   command=self.tree.yview, style="Dk.Vertical.TScrollbar")
        hsb = ttk.Scrollbar(tf, orient=tk.HORIZONTAL, command=self.tree.xview, style="Dk.Horizontal.TScrollbar")
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tf.grid_rowconfigure(0, weight=1)
        tf.grid_columnconfigure(0, weight=1)

        self._reconfigure_tree_tags()

        self.tree.bind("<Double-1>",         self._on_dblclick)
        self.tree.bind("<<TreeviewSelect>>", self._on_tree_select)
        self.tree.bind("<Key>",              self._on_key_press)
        self.tree.bind("<Button-3>",         self._show_context_menu)
        self.tree.bind("<Control-MouseWheel>", lambda e: self._zoom(1 if e.delta > 0 else -1))
        self.tree.bind("<Control-c>",         lambda e: self._copy_selected_paths())
        self.tree.bind("<Control-C>",         lambda e: self._copy_selected_paths())

    # ─────────────────────────────────────────────────────────────────
    # Right sidebar
    # ─────────────────────────────────────────────────────────────────
    def _build_right_sidebar(self, parent):
        tab_bar = tk.Frame(parent, bg=DARK["card"], height=36)
        tab_bar.pack(fill=tk.X)
        tab_bar.pack_propagate(False)

        self.sidebar_tabs: dict[str, tk.Button] = {}
        self.sidebar_content = tk.Frame(parent, bg=DARK["bg"])
        self.sidebar_content.pack(fill=tk.BOTH, expand=True)

        self.summary_tab = tk.Frame(self.sidebar_content, bg=DARK["bg"])
        self.editor_tab  = tk.Frame(self.sidebar_content, bg=DARK["bg"])
        self.stats_tab   = tk.Frame(self.sidebar_content, bg=DARK["bg"])

        for name, label in [("SUMMARY", "🌅 Workday Shift"), ("EDITOR", "💡 File Inspector"), ("STATS", "📊 Stats & Legend")]:
            btn = tk.Button(tab_bar, text=label, font=(self._family, 9, "bold"),
                            bg=DARK["card"], fg=DARK["text_dim"], relief="flat", bd=0,
                            cursor="hand2", command=lambda n=name: self._switch_sidebar_tab(n))
            btn.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            self.sidebar_tabs[name] = btn

        tk.Frame(parent, bg=DARK["accent"], height=2).pack(fill=tk.X)
        self._switch_sidebar_tab("SUMMARY")
        self._build_summary_tab(self.summary_tab)
        self._build_editor_tab(self.editor_tab)
        self._build_stats_tab(self.stats_tab)

    def _switch_sidebar_tab(self, tab_name: str):
        self.summary_tab.pack_forget()
        self.editor_tab.pack_forget()
        self.stats_tab.pack_forget()
        for name, btn in self.sidebar_tabs.items():
            btn.configure(bg=DARK["surface"] if name == tab_name else DARK["card"],
                          fg=DARK["accent"]   if name == tab_name else DARK["text_dim"])
        if tab_name == "SUMMARY":
            self.summary_tab.pack(fill=tk.BOTH, expand=True)
        elif tab_name == "EDITOR":
            self.editor_tab.pack(fill=tk.BOTH, expand=True)
        else:
            self.stats_tab.pack(fill=tk.BOTH, expand=True)

    # ─────────────────────────────────────────────────────────────────
    # Summary (What's Up) tab
    # ─────────────────────────────────────────────────────────────────
    def _build_summary_tab(self, parent):
        sf = ScrollableFrame(parent, bg=DARK["bg"])
        sf.pack(fill=tk.BOTH, expand=True)
        self.summary_inner = sf.scrollable_frame
        self._refresh_summary_tab()

    def _toggle_edit_fields(self):
        """Collapses or expands the editable metadata inputs in the File Inspector."""
        if hasattr(self, "edit_fields_container"):
            if self.current_editor_idx is not None:
                path = self.rows[self.current_editor_idx].get("Path", "")
                status = self.rows[self.current_editor_idx].get("Status", "")
                if status == "FROZEN" and path not in self.unlocked_frozen_paths:
                    ans = messagebox.askyesno(
                        "❄️ Section Frozen / Locked",
                        f"The file '{Path(path).name}' is marked as FROZEN to protect it from accidental changes.\n\nDo you want to unlock it to edit its registry fields?",
                        icon="warning"
                    )
                    if ans:
                        self.unlocked_frozen_paths.add(path)
                        self.toggle_edit_btn.configure(
                            text="🔓 Section Unlocked — Adjust Registry Fields...",
                            bg=DARK["surface"],
                            fg=DARK["text_dim"]
                        )
                        self.sel_meaning_lbl.config(
                            text="🔓 SECTION UNLOCKED TEMPORARILY\n→ Edit carefully or mark as active when done."
                        )
                    else:
                        return
            
            if self.edit_fields_container.winfo_ismapped():
                self.edit_fields_container.pack_forget()
                status = self.rows[self.current_editor_idx].get("Status", "") if self.current_editor_idx is not None else ""
                path = self.rows[self.current_editor_idx].get("Path", "") if self.current_editor_idx is not None else ""
                if status == "FROZEN" and path not in self.unlocked_frozen_paths:
                    self.toggle_edit_btn.configure(text="🔒 Section Frozen — Click to Unlock...", bg=DARK["surface"], fg=DARK["accent"])
                else:
                    self.toggle_edit_btn.configure(text="⚙️ Adjust Registry Fields...", bg=DARK["surface"], fg=DARK["text_dim"])
            else:
                self.edit_fields_container.pack(fill=tk.X, after=self.toggle_edit_btn, pady=4)
                self.toggle_edit_btn.configure(text="Close Editor ❌", bg=DARK["surface"], fg=DARK["danger"])

    def _toggle_git_sync_card(self):
        """Toggles the collapsible git tools frame."""
        self.git_sync_collapsed = not self.git_sync_collapsed
        self._update_git_sync_card_packing()

    def _update_git_sync_card_packing(self):
        """Updates the visible state of the git tools frame based on preference."""
        if not hasattr(self, "git_sync_card") or not hasattr(self, "btn_toggle_git"):
            return
        if self.git_sync_collapsed:
            self.git_sync_card.pack_forget()
            self.btn_toggle_git.configure(text="🔄 Show Git & Auto-Promote Tools...", bg=DARK["surface"], fg=DARK["text_dim"])
        else:
            self.git_sync_card.pack(fill=tk.X, after=self.btn_toggle_git, pady=(0, 6), padx=4)
            self.btn_toggle_git.configure(text="Hide Git Tools ❌", bg=DARK["surface"], fg=DARK["danger"])

    def _toggle_shortcuts_card(self):
        """Toggles the collapsible shortcuts frame."""
        self.shortcuts_collapsed = not self.shortcuts_collapsed
        self._update_shortcuts_card_packing()

    def _update_shortcuts_card_packing(self):
        """Updates the visible state of the shortcuts frame based on preference."""
        if not hasattr(self, "shortcuts_card") or not hasattr(self, "btn_toggle_shortcuts"):
            return
        if self.shortcuts_collapsed:
            self.shortcuts_card.pack_forget()
            self.btn_toggle_shortcuts.configure(text="⌨️ Show Keyboard Shortcuts...", bg=DARK["surface"], fg=DARK["text_dim"])
        else:
            self.shortcuts_card.pack(fill=tk.X, after=self.btn_toggle_shortcuts, pady=(0, 6), padx=4)
            self.btn_toggle_shortcuts.configure(text="Hide Shortcuts ❌", bg=DARK["surface"], fg=DARK["danger"])

    def _get_time_ago_str(self, abs_path):
        """Calculates a friendly ADHD elapsed time string for a modified file."""
        import time
        try:
            mtime = abs_path.stat().st_mtime
            elapsed = int(time.time() - mtime)
            if elapsed < 60:
                return "just now"
            elif elapsed < 3600:
                return f"{elapsed // 60}m ago"
            elif elapsed < 86400:
                return f"{elapsed // 3600}h ago"
            else:
                import datetime
                dt = datetime.datetime.fromtimestamp(mtime)
                return dt.strftime("%b %d")
        except Exception:
            return ""

    def _update_workday_timer_ui(self):
        """Updates the real-time countdown clock in the Workday summary card."""
        if not hasattr(self, "lbl_workday_status"):
            return
        if self.workday_active and self.workday_start_time:
            import datetime
            try:
                start = datetime.datetime.fromisoformat(self.workday_start_time)
                elapsed = datetime.datetime.now() - start
                total_secs = self.workday_duration_mins * 60
                remaining_secs = int(total_secs - elapsed.total_seconds())
                
                if remaining_secs > 0:
                    mins = remaining_secs // 60
                    secs = remaining_secs % 60
                    time_str = f"{mins:02d}:{secs:02d}"
                    self.lbl_workday_status.configure(
                        text=f"🟢 Active Shift: {time_str} remaining",
                        fg=DARK["success"]
                    )
                    self.btn_punch.configure(text="⏱️ Punch Out", bg=DARK["danger"])
                else:
                    self.lbl_workday_status.configure(
                        text="🔴 Shift Finished! (Grace Period)",
                        fg=DARK["warn"]
                    )
                    self.btn_punch.configure(text="⏱️ Punch Out", bg=DARK["danger"])
            except Exception:
                pass
        else:
            self.lbl_workday_status.configure(
                text="🔴 Clocked Out",
                fg=DARK["text_dim"]
            )
            self.btn_punch.configure(text="⏱️ Punch In (60m)", bg=DARK["accent"])
            
        # Run every second
        self.root.after(1000, self._update_workday_timer_ui)

    def _workday_punch_in_out(self):
        """Punches the workday shift clock in or out."""
        if self.workday_active:
            self.workday_active = False
            self.workday_start_time = None
            self._persist_config()
            self._refresh_summary_tab()
            self.sb_left.config(text="⏱️ Workday shift stopped.")
        else:
            import datetime
            self.workday_active = True
            self.workday_start_time = datetime.datetime.now().isoformat()
            
            # Capture start time for all active items that are CONSTRUCTION or NEXT
            self.workday_task_start_times = {}
            for i, d in enumerate(self.rows):
                if d.get("Status") in ("CONSTRUCTION", "NEXT"):
                    p = d.get("Path", "")
                    if p:
                        self.workday_task_start_times[p] = datetime.datetime.now().isoformat()
            
            self._persist_config()
            self._refresh_summary_tab()
            self.sb_left.config(text="⏱️ Workday shift started! Clock is running.")

    def _log_workday_task_research(self, path, duration_secs):
        """Appends a task completion research entry to WORKDAY_RESEARCH.md in the project root."""
        if not self.linked_project_root or not os.path.isdir(self.linked_project_root):
            return
        
        research_file = Path(self.linked_project_root) / "WORKDAY_RESEARCH.md"
        import datetime
        now = datetime.datetime.now()
        timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
        mins = duration_secs // 60
        secs = duration_secs % 60
        duration_str = f"{mins}m {secs}s"
        
        content = ""
        if research_file.exists():
            try:
                content = research_file.read_text(encoding="utf-8")
            except Exception:
                pass
        
        if not content.strip():
            content = f"""# ⏱️ Workday Task Completion Research Log

This file compiles professional research data on AI engineering task completion speed and progress over timed workday shifts.

| Timestamp | Component File | Time to Build | Status | Notes |
| :--- | :--- | :--- | :--- | :--- |
"""
        
        content += f"| {timestamp} | `{path}` | {duration_str} | ✅ Shipped | Bounded workday task completion. |\n"
        
        try:
            research_file.write_text(content, encoding="utf-8")
            self.sb_left.config(text=f"📊 Research logged: {path} built in {duration_str}!")
        except Exception as e:
            print(f"Failed to write workday research: {e}")

    def _trigger_status_change_hook(self, idx, old_status, new_status):
        """Hook triggered whenever an item's status changes to record research timing."""
        d = self.rows[idx]
        path = d.get("Path", "")
        if not path:
            return
            
        import datetime
        now = datetime.datetime.now()
        
        # 1. Track task start times
        if new_status in ("CONSTRUCTION", "NEXT") and old_status not in ("CONSTRUCTION", "NEXT"):
            self.workday_task_start_times[path] = now.isoformat()
            self._persist_config()
            
        # 2. Track task completion & log research
        if new_status in ("SHIP_READY", "ACTIVE_V1", "FROZEN") and old_status in ("CONSTRUCTION", "NEXT", "PLACEHOLDER"):
            if self.workday_active and self.workday_start_time:
                start_iso = self.workday_task_start_times.get(path)
                if not start_iso:
                    start_iso = self.workday_start_time
                
                try:
                    start = datetime.datetime.fromisoformat(start_iso)
                    duration_secs = int((now - start).total_seconds())
                    if duration_secs < 1:
                        duration_secs = 1
                        
                    entry = {
                        "path": path,
                        "duration_secs": duration_secs,
                        "timestamp": now.isoformat()
                    }
                    self.workday_completed_tasks.append(entry)
                    
                    if path in self.workday_task_start_times:
                        del self.workday_task_start_times[path]
                        
                    self._persist_config()
                    self._log_workday_task_research(path, duration_secs)
                    self._refresh_summary_tab()
                except Exception as e:
                    print(f"Error in status change hook: {e}")

    def _format_size(self, size_in_bytes):
        if size_in_bytes < 1024:
            return f"{size_in_bytes} B"
        elif size_in_bytes < 1024 * 1024:
            return f"{size_in_bytes/1024:.1f} KB"
        else:
            return f"{size_in_bytes/(1024*1024):.1f} MB"

    def _get_recently_committed_files(self):
        """Runs git log in linked project to find recently changed files."""
        if not self.linked_project_root or not os.path.isdir(self.linked_project_root):
            return []
        git_dir = Path(self.linked_project_root) / ".git"
        if not git_dir.exists():
            return []
        
        import subprocess
        try:
            res = subprocess.run(
                ["git", "log", "-n", "10", "--name-only", "--pretty=format:"],
                cwd=self.linked_project_root,
                capture_output=True,
                text=True,
                check=True
            )
            files = []
            seen = set()
            for line in res.stdout.splitlines():
                f = line.strip().replace("\\", "/")
                if not f or f in seen:
                    continue
                seen.add(f)
                files.append(f)
            return files
        except Exception as e:
            print(f"Failed to get git log: {e}")
            return []

    def _select_and_open_file(self, rel_path):
        """Focuses/selects the file in the Treeview and opens it in the editor."""
        idx = None
        for i, d in enumerate(self.rows):
            if d.get("Path") == rel_path:
                idx = i
                break
        
        if idx is not None:
            iid = str(idx)
            if self.tree.exists(iid):
                self.tree.selection_set(iid)
                self.tree.focus(iid)
                self.tree.see(iid)
                self._load_detail_panel(idx)
        
        self._open_file_in_editor(rel_path)

    def _refresh_summary_tab(self):
        if not hasattr(self, "summary_inner"):
            return
        
        # Clear existing content
        for w in self.summary_inner.winfo_children():
            w.destroy()
            
        if not self.rows:
            # Show empty placeholder
            tk.Label(self.summary_inner, text="🌅", font=("Segoe UI Emoji", 48),
                     bg=DARK["bg"], fg=DARK["text_dim"]).pack(expand=True, pady=(90, 8))
            tk.Label(self.summary_inner, text="No project skeleton loaded yet.\nOpen a workbook to see your ADHD morning summary!",
                     font=(self._family, 9), bg=DARK["bg"], fg=DARK["text_dim"], justify=tk.CENTER
                     ).pack(expand=True, pady=(0, 90))
            return
            
        # Filter rows by status to find what's left
        todo_construction = []
        todo_next = []
        todo_stubs = []
        todo_gates = []
        todo_roadmap = []
        
        for i, d in enumerate(self.rows):
            if d.get("Type", "").lower() in ("folder", "repo"):
                continue
            status = d.get("Status", "")
            if status == "CONSTRUCTION":
                todo_construction.append((i, d))
            elif status == "NEXT":
                todo_next.append((i, d))
            elif status == "PLACEHOLDER":
                todo_stubs.append((i, d))
            elif status == "HUMAN_GATE":
                todo_gates.append((i, d))
            elif status == "ROADMAP":
                todo_roadmap.append((i, d))

        # Helper to draw a beautiful collapsible header
        def _create_collapsible_card(key, title_text, icon_prefix=""):
            card = tk.Frame(self.summary_inner, bg=DARK["card"], highlightthickness=1,
                            highlightbackground=DARK["border"])
            card.pack(fill=tk.X, pady=(4, 6), padx=4)
            
            header_f = tk.Frame(card, bg=DARK["card"])
            header_f.pack(fill=tk.X)
            
            is_collapsed = self.collapsed_sections.get(key, False)
            arrow_char = "▶" if is_collapsed else "▼"
            
            lbl_arrow = tk.Label(header_f, text=f" {arrow_char} ", font=(self._mono, 10, "bold"),
                                 bg=DARK["card"], fg=DARK["accent"])
            lbl_arrow.pack(side=tk.LEFT, padx=(8, 2), pady=8)
            
            lbl_title = tk.Label(header_f, text=f"{icon_prefix}  {title_text}", font=(self._family, 10, "bold"),
                                 bg=DARK["card"], fg=DARK["text"])
            lbl_title.pack(side=tk.LEFT, pady=8)
            
            # Clicking anywhere on the header toggles it
            def _toggle(e=None):
                self.collapsed_sections[key] = not self.collapsed_sections[key]
                self._refresh_summary_tab()
                
            header_f.bind("<Button-1>", _toggle)
            lbl_arrow.bind("<Button-1>", _toggle)
            lbl_title.bind("<Button-1>", _toggle)
            
            header_f.bind("<Enter>", lambda e: [lbl_title.configure(fg=DARK["accent"]), header_f.configure(cursor="hand2")])
            header_f.bind("<Leave>", lambda e: [lbl_title.configure(fg=DARK["text"])])
            
            if is_collapsed:
                return None
                
            body = tk.Frame(card, bg=DARK["card"])
            body.pack(fill=tk.X, padx=12, pady=(0, 10))
            return body

        # --------------------------------------------------------
        # 0. CARD: ⏱️ WORKDAY SHIFT & TIMING stopwatches
        # --------------------------------------------------------
        workday_body = _create_collapsible_card("workday", "Workday Shift & Research", "⏱️")
        if workday_body:
            lbl_status = tk.Label(workday_body, text="🔴 Clocked Out" if not self.workday_active else "🟢 Active Session Running", 
                                  font=(self._family, 9, "bold"), bg=DARK["card"], fg=DARK["text_dim"])
            lbl_status.pack(anchor=tk.W, pady=2)
            
            btn_f = tk.Frame(workday_body, bg=DARK["card"])
            btn_f.pack(fill=tk.X, pady=(4, 8))
            
            btn_punch = tk.Button(btn_f, text="⏱️ Punch Out" if self.workday_active else "⏱️ Punch In (60m)", 
                                  font=(self._family, 8, "bold"),
                                  bg=DARK["accent"], fg="#07101E", relief="flat", bd=0,
                                  padx=12, pady=4, cursor="hand2",
                                  command=self._workday_punch_in_out)
            btn_punch.pack(side=tk.LEFT)
            
            # If there are completed tasks in this workday, show them as a research log
            if self.workday_completed_tasks:
                tk.Label(workday_body, text="📊 Current Shift Research Log:", font=(self._family, 8, "bold"),
                         bg=DARK["card"], fg=DARK["text"]).pack(anchor=tk.W, pady=(8, 2))
                         
                log_f = tk.Frame(workday_body, bg=DARK["card"])
                log_f.pack(fill=tk.X, pady=(2, 4))
                
                for task in self.workday_completed_tasks[-6:]: # Show up to 6 completed tasks when expanded
                    p = task.get("path", "")
                    dur = task.get("duration_secs", 0)
                    mins = dur // 60
                    secs = dur % 60
                    dur_str = f"{mins}m {secs}s"
                    name = Path(p).name
                    
                    row_t = tk.Frame(log_f, bg=DARK["card"])
                    row_t.pack(fill=tk.X, pady=2)
                    
                    tk.Label(row_t, text=f"✓  {name}", font=(self._family, 8),
                             bg=DARK["card"], fg=DARK["success"]).pack(side=tk.LEFT)
                             
                    tk.Label(row_t, text=f"⏱️ {dur_str} (Logged)", font=(self._family, 8, "italic"),
                             bg=DARK["card"], fg=DARK["text_dim"]).pack(side=tk.RIGHT)

        # --------------------------------------------------------
        # 0.5. CARD: 🤖 MASTER AI GAMEPLAN COPIER (VIBE CODING ASSISTANT)
        # --------------------------------------------------------
        copier_body = _create_collapsible_card("copier", "Master AI Gameplan Copier", "🤖")
        if copier_body:
            tk.Label(copier_body, text="Instantly copy high-density goals, roadmap, active rules, and file lists to paste straight into your AI coding agent chat.",
                     font=(self._family, 8), bg=DARK["card"], fg=DARK["text_dim"], justify=tk.LEFT, wraplength=360).pack(anchor=tk.W, pady=(2, 8))
            
            btn_copy = tk.Button(copier_body, text="📋 Copy Master AI Gameplan Brief", font=(self._family, 9, "bold"),
                                 bg=DARK["accent"], fg="#07101E", relief="flat", bd=0,
                                 padx=16, pady=6, cursor="hand2",
                                 command=self._copy_ai_brief)
            btn_copy.pack(fill=tk.X, pady=2)

        # --------------------------------------------------------
        # 1. CARD: 🎯 DO THIS FIRST (ACTIONABLE RECOMMENDATION)
        # --------------------------------------------------------
        recommend_body = _create_collapsible_card("recommend", "Do This First!", "🎯")
        if recommend_body:
            recommended = []
            rec_reason = ""
            if todo_construction:
                recommended = todo_construction[:2]
                rec_reason = "Already under construction! Keep coding this room."
            elif todo_next:
                recommended = todo_next[:2]
                rec_reason = "Approved and queued for work. Let's build this room next!"
            elif todo_stubs:
                recommended = todo_stubs[:2]
                rec_reason = "This is an empty placeholder/stub file. Let's start building it!"
            
            if recommended:
                for idx, d in recommended:
                    p = d.get("Path", "")
                    status = d.get("Status", "")
                    icon = STATUS_ICONS.get(status, "📄")
                    
                    row_f = tk.Frame(recommend_body, bg=DARK["card"])
                    row_f.pack(fill=tk.X, pady=3)
                    
                    lbl_file = tk.Label(row_f, text=f"{icon} {p}", font=(self._family, 9, "bold"),
                                        bg=DARK["card"], fg=DARK["accent"], anchor=tk.W, cursor="hand2")
                    lbl_file.pack(side=tk.LEFT, fill=tk.X, expand=True)
                    
                    lbl_file.bind("<Enter>", lambda e, lbl=lbl_file: lbl.configure(font=(self._family, 9, "bold", "underline")))
                    lbl_file.bind("<Leave>", lambda e, lbl=lbl_file: lbl.configure(font=(self._family, 9, "bold")))
                    lbl_file.bind("<Button-1>", lambda e, path=p: self._select_and_open_file(path))
                    
                tk.Label(recommend_body, text=f"💡 Why? {rec_reason}", font=(self._family, 8, "italic"),
                         bg=DARK["card"], fg=DARK["text_dim"], justify=tk.LEFT, wraplength=360).pack(anchor=tk.W, pady=(6, 2))
            else:
                tk.Label(recommend_body, text="No pending tasks found! Everything is verified and clean. 🎉", font=(self._family, 8, "bold"),
                         bg=DARK["card"], fg=DARK["success"]).pack(anchor=tk.W, pady=4)

        # --------------------------------------------------------
        # 2. CARD: 🕒 RECENTLY WORKED ON (EXPANDED TO BE LARGER)
        # --------------------------------------------------------
        recent_body = _create_collapsible_card("recent", "Recently Worked On", "🕒")
        if recent_body:
            project_root = Path(self.linked_project_root) if self.linked_project_root else (Path(self.xlsx_path).parent if self.xlsx_path else Path("."))
            
            # Sort ALL existing files in registry by modification time to find the exact order
            recent_files = []
            if project_root:
                for i, d in enumerate(self.rows):
                    if d.get("Type", "").lower() in ("folder", "repo"):
                        continue
                    p = d.get("Path", "")
                    abs_path = project_root / p
                    if abs_path.is_file():
                        try:
                            mtime = abs_path.stat().st_mtime
                            sz = abs_path.stat().st_size
                            recent_files.append((i, d, mtime, sz))
                        except Exception:
                            pass
            
            # Sort descending (most recently modified first!)
            recent_files.sort(key=lambda x: x[2], reverse=True)
            
            # Temporal clustering into runs
            runs = []
            if recent_files:
                current_run = [recent_files[0]]
                for item in recent_files[1:]:
                    prev_item = current_run[-1]
                    prev_mtime = prev_item[2]
                    curr_mtime = item[2]
                    # 15 minutes (900 seconds) gap threshold defines a new coding run
                    if abs(prev_mtime - curr_mtime) <= 900:
                        current_run.append(item)
                    else:
                        runs.append(current_run)
                        current_run = [item]
                if current_run:
                    runs.append(current_run)
            
            import time
            if runs:
                # Show up to 5 temporal runs (expanded) instead of 3 to make the list much bigger!
                for r_idx, run in enumerate(runs[:5]):
                    max_mtime = run[0][2]
                    elapsed = int(time.time() - max_mtime)
                    
                    if elapsed < 900:
                        run_title = "🔥 Active Coding Session (Just Now)"
                        title_color = DARK["accent"]
                    elif elapsed < 3600:
                        run_title = f"💻 Coding Session — {elapsed // 60}m ago"
                        title_color = DARK["text"]
                    elif elapsed < 86400:
                        run_title = f"💻 Coding Session — {elapsed // 3600}h ago"
                        title_color = DARK["text"]
                    else:
                        import datetime
                        dt = datetime.datetime.fromtimestamp(max_mtime)
                        run_title = f"📅 Coding Session — {dt.strftime('%b %d')}"
                        title_color = DARK["text_dim"]
                        
                    tk.Label(recent_body, text=f"{run_title} ({len(run)} files):", font=(self._family, 8, "bold"),
                             bg=DARK["card"], fg=title_color).pack(anchor=tk.W, pady=(6, 2))
                             
                    # List files in this run (up to 10 files per run when expanded) instead of 4!
                    for idx, d, mtime, sz in run[:10]:
                        p = d.get("Path", "")
                        status = d.get("Status", "")
                        icon = STATUS_ICONS.get(status, "📄")
                        size_str = f" ({self._format_size(sz)})"
                        time_ago = self._get_time_ago_str(project_root / p)
                        time_desc = f" — {time_ago}" if time_ago else ""
                        
                        is_mod = p in self.git_changed_files
                        mod_fg = DARK["accent"] if is_mod else DARK["text"]
                        mod_badge = " [MODIFIED]" if is_mod else ""
                        
                        row_f = tk.Frame(recent_body, bg=DARK["card"])
                        row_f.pack(fill=tk.X, padx=8, pady=2)
                        
                        lbl_file = tk.Label(row_f, text=f"•  {icon} {p}{size_str}{time_desc}{mod_badge}", font=(self._family, 8),
                                            bg=DARK["card"], fg=mod_fg, anchor=tk.W, cursor="hand2")
                        lbl_file.pack(side=tk.LEFT, fill=tk.X, expand=True)
                        
                        lbl_file.bind("<Enter>", lambda e, lbl=lbl_file: lbl.configure(font=(self._family, 8, "underline"), fg=DARK["accent"]))
                        lbl_file.bind("<Leave>", lambda e, lbl=lbl_file, cfg=mod_fg: lbl.configure(font=(self._family, 8), fg=cfg))
                        lbl_file.bind("<Button-1>", lambda e, path=p: self._select_and_open_file(path))
                        
                    if len(run) > 10:
                        tk.Label(recent_body, text=f"  ...and {len(run) - 10} more files in this session.", font=(self._family, 7, "italic"),
                                 bg=DARK["card"], fg=DARK["text_dim"]).pack(anchor=tk.W, padx=12, pady=(1, 4))
            else:
                tk.Label(recent_body, text="No modified files found on disk.", font=(self._family, 8, "italic"),
                         bg=DARK["card"], fg=DARK["text_dim"]).pack(anchor=tk.W, pady=(4, 6))

        # --------------------------------------------------------
        # 3. CARD: ⏳ CURRENT V1 SKELETON TASKS (EXPANDED TO BE LARGER)
        # --------------------------------------------------------
        todo_body = _create_collapsible_card("todo", "Current V1 Tasks To Do", "⏳")
        if todo_body:
            tk.Frame(todo_body, bg=DARK["border"], height=1).pack(fill=tk.X, pady=2)
            
            has_any_v1_todo = False
            categories = [
                ("🔨 Active Builds (CONSTRUCTION)", todo_construction, DARK["accent"]),
                ("⏭ Queued Up Next (NEXT)", todo_next, self.status_colors.get("NEXT", DARK["text"])),
                ("📋 Skeletons & Placeholders (PLACEHOLDER)", todo_stubs, self.status_colors.get("PLACEHOLDER", DARK["text"])),
            ]
            
            for cat_name, items, cat_color in categories:
                if items:
                    has_any_v1_todo = True
                    tk.Label(todo_body, text=f"{cat_name} ({len(items)}):", font=(self._family, 8, "bold"),
                             bg=DARK["card"], fg=cat_color).pack(anchor=tk.W, pady=(6, 2))
                    # Render up to 15 items (expanded) in this category to prevent overwhelm but allow long list scanning!
                    for idx, d in items[:15]:
                        p = d.get("Path", "")
                        row_f = tk.Frame(todo_body, bg=DARK["card"])
                        row_f.pack(fill=tk.X, padx=8, pady=2)
                        
                        icon = STATUS_ICONS.get(d.get("Status", ""), "📄")
                        lbl_file = tk.Label(row_f, text=f"• {icon} {p}", font=(self._family, 8),
                                            bg=DARK["card"], fg=DARK["text"], anchor=tk.W, cursor="hand2")
                        lbl_file.pack(side=tk.LEFT, fill=tk.X, expand=True)
                        
                        lbl_file.bind("<Enter>", lambda e, lbl=lbl_file: lbl.configure(font=(self._family, 8, "underline"), fg=DARK["accent"]))
                        lbl_file.bind("<Leave>", lambda e, lbl=lbl_file: lbl.configure(font=(self._family, 8), fg=DARK["text"]))
                        lbl_file.bind("<Button-1>", lambda e, path=p: self._select_and_open_file(path))
                    
                    if len(items) > 15:
                        tk.Label(todo_body, text=f"  ...and {len(items) - 15} more files.", font=(self._family, 7, "italic"),
                                 bg=DARK["card"], fg=DARK["text_dim"]).pack(anchor=tk.W, padx=12, pady=(1, 4))
                                 
            if not has_any_v1_todo:
                tk.Label(todo_body, text="No pending tasks remaining in V1! 🎉", font=(self._family, 8, "bold"),
                         bg=DARK["card"], fg=DARK["success"]).pack(anchor=tk.W, pady=(4, 6))

        # --------------------------------------------------------
        # 4. CARD: ✨ NEXT HORIZONS & POLISHING (POST-V1 OR BOTTLENECKS)
        # --------------------------------------------------------
        horizon_body = _create_collapsible_card("horizon", "Next Horizons & Polishing", "✨")
        if horizon_body:
            tk.Frame(horizon_body, bg=DARK["border"], height=1).pack(fill=tk.X, pady=2)
            
            # Check if there's any active V1 todo left
            v1_todo_exists = len(todo_construction) > 0 or len(todo_next) > 0 or len(todo_stubs) > 0
            
            if not v1_todo_exists:
                tk.Label(horizon_body, text="🎉 All V1 Rooms Built! Keep Working & Polishing:",
                         font=(self._family, 9, "bold"), bg=DARK["card"], fg=DARK["success"]).pack(anchor=tk.W, pady=(4, 2))
                
                polishing_tips = [
                    "❄️ Freeze Completed Rooms: Set completed rooms to 🔒 FROZEN to lock them.",
                    "🗺 Promote Future Features: Select any 🗺 ROADMAP file to queue it next.",
                    "🧪 Regression check: Rerun automated unit tests.",
                    "🎨 UI inspection: Inspect visual layouts for alignment and padding."
                ]
                for tip in polishing_tips:
                    tk.Label(horizon_body, text=f"• {tip}", font=(self._family, 8),
                             bg=DARK["card"], fg=DARK["text_dim"], justify=tk.LEFT, anchor=tk.W, wraplength=360
                             ).pack(anchor=tk.W, padx=8, pady=1)
            else:
                # Suggestions for future
                tk.Label(horizon_body, text="Ready for V2? Promote a 🗺 FUTURE file once V1 is done.",
                         font=(self._family, 8, "italic"), bg=DARK["card"], fg=DARK["text_dim"]).pack(anchor=tk.W, pady=(4, 4))

            # Gate Bottlenecks (HUMAN_GATE)
            if todo_gates:
                tk.Label(horizon_body, text="🚧 Blocked at Gate (HUMAN_GATE):", font=(self._family, 8, "bold"),
                         bg=DARK["card"], fg=self.status_colors.get("HUMAN_GATE", DARK["text"])).pack(anchor=tk.W, pady=(6, 2))
                for idx, d in todo_gates[:5]: # Show up to 5 gates
                    p = d.get("Path", "")
                    row_f = tk.Frame(horizon_body, bg=DARK["card"])
                    row_f.pack(fill=tk.X, padx=8, pady=2)
                    lbl_file = tk.Label(row_f, text=f"• 🚧 {p}", font=(self._family, 8),
                                        bg=DARK["card"], fg=DARK["text"], anchor=tk.W, cursor="hand2")
                    lbl_file.pack(side=tk.LEFT, fill=tk.X, expand=True)
                    lbl_file.bind("<Enter>", lambda e, lbl=lbl_file: lbl.configure(font=(self._family, 8, "underline"), fg=DARK["accent"]))
                    lbl_file.bind("<Leave>", lambda e, lbl=lbl_file: lbl.configure(font=(self._family, 8), fg=DARK["text"]))
                    lbl_file.bind("<Button-1>", lambda e, path=p: self._select_and_open_file(path))
                if len(todo_gates) > 5:
                    tk.Label(horizon_body, text=f"  ...and {len(todo_gates) - 5} more blocked files.", font=(self._family, 7, "italic"),
                             bg=DARK["card"], fg=DARK["text_dim"]).pack(anchor=tk.W, padx=12, pady=(1, 4))
                             
            # Roadmap (ROADMAP) features ready for promotion
            if todo_roadmap:
                tk.Label(horizon_body, text="🗺 Future Roadmap (Promote to Queue):", font=(self._family, 8, "bold"),
                         bg=DARK["card"], fg=self.status_colors.get("ROADMAP", DARK["text"])).pack(anchor=tk.W, pady=(6, 2))
                for idx, d in todo_roadmap[:5]: # Show up to 5 roadmap files
                    p = d.get("Path", "")
                    row_f = tk.Frame(horizon_body, bg=DARK["card"])
                    row_f.pack(fill=tk.X, padx=8, pady=2)
                    lbl_file = tk.Label(row_f, text=f"• 🗺 {p}", font=(self._family, 8),
                                        bg=DARK["card"], fg=DARK["text"], anchor=tk.W, cursor="hand2")
                    lbl_file.pack(side=tk.LEFT, fill=tk.X, expand=True)
                    lbl_file.bind("<Enter>", lambda e, lbl=lbl_file: lbl.configure(font=(self._family, 8, "underline"), fg=DARK["accent"]))
                    lbl_file.bind("<Leave>", lambda e, lbl=lbl_file: lbl.configure(font=(self._family, 8), fg=DARK["text"]))
                    lbl_file.bind("<Button-1>", lambda e, path=p: self._select_and_open_file(path))
                if len(todo_roadmap) > 5:
                    tk.Label(horizon_body, text=f"  ...and {len(todo_roadmap) - 5} more future roadmap files.", font=(self._family, 7, "italic"),
                             bg=DARK["card"], fg=DARK["text_dim"]).pack(anchor=tk.W, padx=12, pady=(1, 4))

        tk.Frame(horizon_body, bg=DARK["card"], height=6).pack()

    def _copy_ai_brief(self):
        if not self.rows:
            messagebox.showinfo("Nothing to copy", "No workspace loaded.")
            return
            
        proj_name = Path(self.linked_project_root).name if getattr(self, "repository_mode", False) else (Path(self.xlsx_path).stem if self.xlsx_path else "Workspace")
        branch_desc = f"Branch: {self.current_git_branch or 'main'}"
        
        construction_files = []
        next_files = []
        placeholder_files = []
        
        for d in self.rows:
            status = d.get("Status", "")
            path = d.get("Path", "")
            purpose = d.get("Purpose", "")
            if not path or d.get("Type", "").lower() in ("folder", "repo"):
                continue
            if status == "CONSTRUCTION":
                construction_files.append(f"  • `{path}` — {purpose}")
            elif status == "NEXT":
                next_files.append(f"  • `{path}` — {purpose}")
            elif status == "PLACEHOLDER":
                placeholder_files.append(f"  • `{path}` — {purpose}")
                
        prompt = []
        prompt.append(f"# 🪐 Unified AI Control Plane (UACP) Master Gameplan Brief")
        prompt.append(f"You are the senior AI pair-programmer working alongside James (the Vibe Coder) in **{proj_name}**.")
        prompt.append(f"Your workspace path: `{self.linked_project_root or 'C:\\Projects\\' + proj_name}`")
        prompt.append(f"Current Git State: {branch_desc}")
        prompt.append("")
        
        prompt.append("## 🎯 Active Tasks in Progress")
        if construction_files:
            prompt.append("### 🔨 CURRENTLY UNDER CONSTRUCTION (Your main focus):")
            prompt.extend(construction_files)
        else:
            prompt.append("No files are currently marked under active CONSTRUCTION.")
            
        if next_files:
            prompt.append("\n### ⏭ QUEUED UP NEXT:")
            prompt.extend(next_files)
            
        if placeholder_files:
            prompt.append("\n### 📋 BARE SKELETONS / PLACEHOLDERS (Ready to build):")
            prompt.extend(placeholder_files[:12])
            if len(placeholder_files) > 12:
                prompt.append(f"  ...and {len(placeholder_files) - 12} more placeholders.")
                
        prompt.append("")
        prompt.append("## 🛡️ VIBE CODING OPERATING BOUNDARIES")
        prompt.append("1. **Complete Autonomy:** Take 100% technical ownership of files inside active scope. Do not ask James for architectural green lights.")
        prompt.append("2. **Surgical Precision:** Do NOT rewrite entire large files. Only touch active files listed above.")
        prompt.append("3. **Decoupled Workday Clutter:** The workday timetracking stopwatch and shift timer reside strictly in the external desktop app and agent memory skills. Never contaminate the core CLI codebase with workday timers.")
        prompt.append("4. **Rule Integrity:** Before any commit or closeout, check that you have fully updated the handoff documents (`COMPACT_HANDOFF.md` or local rules).")
        prompt.append("\nLet's get straight to work on implementing the current active files! Show me your plan for the next action.")
        
        brief_text = "\n".join(prompt)
        
        self.root.clipboard_clear()
        self.root.clipboard_append(brief_text)
        
        messagebox.showinfo("Success", "📋 Master AI Gameplan Brief copied to clipboard successfully!\n\nPaste this directly into your next AI chat turn to start vibe coding with 100% context alignment.")

    # ─────────────────────────────────────────────────────────────────
    # Editor tab
    # ─────────────────────────────────────────────────────────────────
    def _build_editor_tab(self, parent):
        self.editor_container = tk.Frame(parent, bg=DARK["bg"])
        self.editor_container.pack(fill=tk.BOTH, expand=True)

        # Placeholder
        self.editor_placeholder = tk.Frame(self.editor_container, bg=DARK["card"],
                                           highlightthickness=1, highlightbackground=DARK["border"])
        self.editor_placeholder.pack(fill=tk.BOTH, expand=True)
        tk.Label(self.editor_placeholder, text="🤖", font=("Segoe UI Emoji", 48),
                 bg=DARK["card"], fg=DARK["text_dim"]).pack(expand=True, pady=(90, 8), anchor=tk.S)
        tk.Label(self.editor_placeholder,
                 text="Select any file or folder in the tree\nto inspect, edit status, or generate AI prompts.",
                 font=(self._family, 9), bg=DARK["card"], fg=DARK["text_dim"], justify=tk.CENTER
                 ).pack(expand=True, pady=(0, 90), anchor=tk.N)

        # Scrollable editor fields
        self.editor_fields_frame = tk.Frame(self.editor_container, bg=DARK["bg"])

        sf = ScrollableFrame(self.editor_fields_frame, bg=DARK["bg"])
        sf.pack(fill=tk.BOTH, expand=True)
        inner = sf.scrollable_frame

        card = tk.Frame(inner, bg=DARK["card"], highlightthickness=1,
                        highlightbackground=DARK["border"])
        card.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)

        # Header
        hdr = tk.Frame(card, bg=DARK["card"])
        hdr.pack(fill=tk.X, padx=12, pady=(12, 6))

        self.sel_icon_lbl = tk.Label(hdr, text="📄", font=("Segoe UI Emoji", 14),
                                     bg=DARK["card"], fg=DARK["accent"])
        self.sel_icon_lbl.pack(side=tk.LEFT, padx=(0, 4))

        self.sel_path_lbl = tk.Label(hdr, text="file.py",
                                     font=(self._family, 10, "bold"),
                                     bg=DARK["card"], fg=DARK["text"], anchor=tk.W)
        self.sel_path_lbl.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.sel_status_badge = tk.Label(hdr, text="",
                                         font=(self._family, 8, "bold"),
                                         bg=DARK["card"], fg="#0A0A0A", padx=6, pady=2)
        self.sel_status_badge.pack(side=tk.RIGHT)

        # Status meaning explainer
        self.sel_meaning_lbl = tk.Label(card, text="",
                                        font=(self._family, 8, "italic"),
                                        bg=DARK["card"], fg=DARK["text_dim"], anchor=tk.W, wraplength=340)
        self.sel_meaning_lbl.pack(fill=tk.X, padx=12, pady=(0, 4))

        tk.Frame(card, bg=DARK["border"], height=1).pack(fill=tk.X, padx=12, pady=(0, 8))

        # Read-Only Insight Frame (Default View)
        self.editor_read_only_frame = tk.Frame(card, bg=DARK["card"])
        self.editor_read_only_frame.pack(fill=tk.X, padx=12, pady=4)
        
        # Meta row: Owner & Phase
        self.lbl_view_meta = tk.Label(self.editor_read_only_frame, text="👤 Owner: Unassigned  •  🎯 Phase: N/A",
                                      font=(self._family, 8, "bold"), bg=DARK["card"], fg=DARK["accent"], anchor=tk.W)
        self.lbl_view_meta.pack(fill=tk.X, pady=(0, 6))
        
        # Purpose block
        tk.Label(self.editor_read_only_frame, text="Purpose:", font=(self._family, 7, "bold"),
                 bg=DARK["card"], fg=DARK["text_dim"], anchor=tk.W).pack(fill=tk.X, pady=(2, 0))
        self.lbl_view_purpose = tk.Label(self.editor_read_only_frame, text="Not specified",
                                         font=(self._family, 8), bg=DARK["surface"], fg=DARK["text"],
                                         anchor=tk.W, justify=tk.LEFT, wraplength=380, padx=6, pady=4)
        self.lbl_view_purpose.pack(fill=tk.X, pady=(0, 6))
        
        # Build Rule block
        tk.Label(self.editor_read_only_frame, text="Build Rule:", font=(self._family, 7, "bold"),
                 bg=DARK["card"], fg=DARK["text_dim"], anchor=tk.W).pack(fill=tk.X, pady=(2, 0))
        self.lbl_view_rule = tk.Label(self.editor_read_only_frame, text="Not specified",
                                      font=(self._mono, 8), bg=DARK["surface"], fg=DARK["accent"],
                                      anchor=tk.W, justify=tk.LEFT, wraplength=380, padx=6, pady=4)
        self.lbl_view_rule.pack(fill=tk.X, pady=(0, 6))
        
        # Notes block
        tk.Label(self.editor_read_only_frame, text="Notes:", font=(self._family, 7, "bold"),
                 bg=DARK["card"], fg=DARK["text_dim"], anchor=tk.W).pack(fill=tk.X, pady=(2, 0))
        self.lbl_view_notes = tk.Label(self.editor_read_only_frame, text="No notes added",
                                       font=(self._family, 8), bg=DARK["surface"], fg=DARK["text"],
                                       anchor=tk.W, justify=tk.LEFT, wraplength=380, padx=6, pady=4)
        self.lbl_view_notes.pack(fill=tk.X, pady=(0, 8))

        # Collapsible Edit Container
        self.edit_fields_container = tk.Frame(card, bg=DARK["card"])
        # (Not packed by default)
        
        # Toggle Button
        self.toggle_edit_btn = tk.Button(card, text="⚙️ Adjust Registry Fields...", font=(self._family, 8, "bold"),
                                         bg=DARK["surface"], fg=DARK["text_dim"], activebackground=DARK["border"],
                                         activeforeground=DARK["text"], relief="flat", bd=0, pady=5, cursor="hand2",
                                         command=self._toggle_edit_fields)
        self.toggle_edit_btn.pack(fill=tk.X, padx=12, pady=4)

        # Fields grid (packed inside Collapsible Container)
        grid = tk.Frame(self.edit_fields_container, bg=DARK["card"])
        grid.pack(fill=tk.X, padx=12, pady=4)
        grid.columnconfigure(1, weight=1)

        self.editor_inputs = {}
        fields = [
            ("Status:",     "Status",     "combobox"),
            ("Owner:",      "Owner",      "entry"),
            ("Phase:",      "Phase",      "entry"),
            ("Purpose:",    "Purpose",    "text_small"),
            ("Build Rule:", "Build Rule", "text_large"),
            ("Notes:",      "Notes",      "text_small"),
        ]
        for r, (label, key, ftype) in enumerate(fields):
            tk.Label(grid, text=label, font=(self._family, 8, "bold"),
                     bg=DARK["card"], fg=DARK["text_dim"], anchor=tk.W
                     ).grid(row=r, column=0, sticky="nw", pady=3, padx=(0, 8))

            if ftype == "combobox":
                var = tk.StringVar()
                cb  = ttk.Combobox(grid, textvariable=var, values=STATUS_ORDER,
                                   state="readonly", font=self.f)
                cb.grid(row=r, column=1, sticky="ew", pady=3)
                cb.bind("<<ComboboxSelected>>", lambda e: self._on_status_combo_change())
                self.editor_inputs[key] = (cb, var)
            elif ftype == "entry":
                var = tk.StringVar()
                ent = tk.Entry(grid, textvariable=var, font=self.f,
                               bg=DARK["surface"], fg=DARK["text"],
                               insertbackground=DARK["text"], relief="flat",
                               highlightthickness=1, highlightcolor=DARK["accent"],
                               highlightbackground=DARK["border"])
                ent.grid(row=r, column=1, sticky="ew", pady=3)
                self.editor_inputs[key] = (ent, var)
            else:
                h = 2 if ftype == "text_small" else 4
                txt = tk.Text(grid, height=h, font=self.f,
                              bg=DARK["surface"], fg=DARK["text"],
                              insertbackground=DARK["text"], relief="flat",
                              highlightthickness=1, highlightcolor=DARK["accent"],
                              highlightbackground=DARK["border"], wrap=tk.WORD)
                txt.grid(row=r, column=1, sticky="ew", pady=3)
                self.editor_inputs[key] = (txt, None)

        self.apply_btn = tk.Button(self.edit_fields_container, text="💾 Apply Changes",
                                   font=(self._family, 9, "bold"),
                                   bg=DARK["accent"], fg="#07101E", relief="flat", bd=0,
                                   activebackground=DARK["accent_hover"], cursor="hand2",
                                   command=self._apply_editor_changes)
        self.apply_btn.pack(fill=tk.X, padx=12, pady=(4, 8))

        # Disk sync
        self.disk_frame = tk.Frame(card, bg=DARK["card"],
                                   highlightthickness=1, highlightbackground=DARK["border"])
        self.disk_frame.pack(fill=tk.X, padx=12, pady=(4, 4))
        tk.Label(self.disk_frame, text="💿 Disk Sync",
                 font=(self._family, 8, "bold"), bg=DARK["card"], fg=DARK["text"]
                 ).pack(anchor=tk.W, padx=8, pady=(6, 2))
        self.editor_disk_lbl = tk.Label(self.disk_frame, text="Select a file",
                                        font=(self._family, 8), bg=DARK["card"], fg=DARK["text_dim"])
        self.editor_disk_lbl.pack(anchor=tk.W, padx=8, pady=(0, 6))
        self.stub_btn        = ttk.Button(self.disk_frame, text="✨ Create Stub on Disk",
                                          style="Ac.TButton", command=self._create_stub_for_selected)
        self.open_editor_btn = ttk.Button(self.disk_frame, text="✏️ Open in Editor",
                                          style="Ac.TButton", command=self._open_selected_in_editor)

        # Git Sync Frame (packed dynamically when git changes are detected for selected file)
        self.git_frame = tk.Frame(card, bg=DARK["card"],
                                  highlightthickness=1, highlightbackground=DARK["border"])
        self.git_title = tk.Label(self.git_frame, text="🔄 Git Changes Detected",
                                  font=(self._family, 8, "bold"), bg=DARK["card"], fg=DARK["accent"])
        self.git_title.pack(anchor=tk.W, padx=8, pady=(6, 2))
        self.git_desc_lbl = tk.Label(self.git_frame, text="File is modified in active project.",
                                     font=(self._family, 8), bg=DARK["card"], fg=DARK["text_dim"])
        self.git_desc_lbl.pack(anchor=tk.W, padx=8, pady=(0, 4))
        btn_f = tk.Frame(self.git_frame, bg=DARK["card"])
        btn_f.pack(fill=tk.X, padx=8, pady=(0, 6))
        self.git_done_btn = tk.Button(btn_f, text="✅ Accept Done",
                                      font=(self._family, 7, "bold"),
                                      bg=DARK["success"], fg="#07101E", relief="flat", bd=0,
                                      padx=10, pady=4, cursor="hand2",
                                      command=self._mark_selected_ship_ready)
        self.git_done_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 2))
        self.git_live_btn = tk.Button(btn_f, text="⚡ Accept Live",
                                      font=(self._family, 7, "bold"),
                                      bg=DARK["accent"], fg="#07101E", relief="flat", bd=0,
                                      padx=10, pady=4, cursor="hand2",
                                      command=self._mark_selected_active_v1)
        self.git_live_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(2, 0))

        # LLM helpers
        llm = tk.Frame(card, bg=DARK["card"])
        llm.pack(fill=tk.X, padx=12, pady=(4, 12))
        tk.Label(llm, text="🤖 AI Pair Programming",
                 font=(self._family, 8, "bold"), bg=DARK["card"], fg=DARK["text"]
                 ).pack(anchor=tk.W, pady=(4, 4))
        self.llm_btn_prompt  = ttk.Button(llm, text="📋 Copy AI Implementation Prompt",
                                          style="Dk.TButton", command=self._copy_prompt_for_selected)
        self.llm_btn_prompt.pack(fill=tk.X, pady=1)
        self.llm_btn_context = ttk.Button(llm, text="📝 Copy Markdown Context Block",
                                          style="Dk.TButton", command=self._copy_context_block_for_selected)
        self.llm_btn_context.pack(fill=tk.X, pady=1)
        self.llm_btn_path    = ttk.Button(llm, text="🔗 Copy Path",
                                          style="Dk.TButton", command=self._copy_path_for_selected)
        self.llm_btn_path.pack(fill=tk.X, pady=1)

        # Smart Code Preview Frame
        self.preview_frame = tk.Frame(card, bg=DARK["card"],
                                      highlightthickness=1, highlightbackground=DARK["border"])
        self.preview_frame.pack(fill=tk.X, padx=12, pady=(4, 12))
        
        # Header for preview
        preview_hdr = tk.Frame(self.preview_frame, bg=DARK["card"])
        preview_hdr.pack(fill=tk.X, padx=8, pady=(6, 4))
        
        self.preview_title = tk.Label(preview_hdr, text="📄 Smart Preview (Code / Git Diff)",
                                      font=(self._family, 8, "bold"), bg=DARK["card"], fg=DARK["accent"])
        self.preview_title.pack(side=tk.LEFT)
        
        self.preview_type_badge = tk.Label(preview_hdr, text="",
                                           font=(self._family, 7, "bold"),
                                           bg=DARK["surface"], fg=DARK["text_dim"], padx=4)
        self.preview_type_badge.pack(side=tk.RIGHT)
        
        # Scrollable Text area for code/diff
        txt_f = tk.Frame(self.preview_frame, bg=DARK["surface"])
        txt_f.pack(fill=tk.X, padx=8, pady=(0, 8))
        
        self.preview_txt = tk.Text(txt_f, height=12, font=self.fm,
                                   bg=DARK["surface"], fg=DARK["text"],
                                   insertbackground=DARK["text"], relief="flat", bd=0,
                                   wrap=tk.NONE)
        
        vsb_p = ttk.Scrollbar(txt_f, orient=tk.VERTICAL, command=self.preview_txt.yview,
                              style="Dk.Vertical.TScrollbar")
        hsb_p = ttk.Scrollbar(self.preview_frame, orient=tk.HORIZONTAL, command=self.preview_txt.xview,
                              style="Dk.Horizontal.TScrollbar")
        
        self.preview_txt.configure(yscrollcommand=vsb_p.set, xscrollcommand=hsb_p.set)
        
        self.preview_txt.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb_p.pack(side=tk.RIGHT, fill=tk.Y)
        hsb_p.pack(fill=tk.X, padx=8, pady=(0, 4))
        
        # Configure tags for diff highlighting
        self.preview_txt.tag_configure("diff_add", foreground="#00E676")
        self.preview_txt.tag_configure("diff_del", foreground="#EF5350")
        self.preview_txt.tag_configure("diff_hdr", foreground="#29B6F6")
        self.preview_txt.tag_configure("info",     foreground=DARK["text_dim"], font=self.fs)

    def _on_status_combo_change(self):
        """Live-update the meaning label when status combobox changes."""
        if hasattr(self, "editor_inputs") and "Status" in self.editor_inputs:
            _, var = self.editor_inputs["Status"]
            status = var.get()
            if status and status in STATUS_WORKFLOW:
                icon_label, meaning, action = STATUS_WORKFLOW[status]
                self.sel_meaning_lbl.config(
                    text=f"{icon_label}  —  {meaning}\n→ {action}")
                bg = self.status_colors.get(status, DARK["card"])
                icon = STATUS_ICONS.get(status, "")
                self.sel_status_badge.config(text=f"{icon} {status}", bg=bg)

    # ─────────────────────────────────────────────────────────────────
    # Stats & Legend tab
    # ─────────────────────────────────────────────────────────────────
    def _build_stats_tab(self, parent):
        sf = ScrollableFrame(parent, bg=DARK["bg"])
        sf.pack(fill=tk.BOTH, expand=True)
        self._stats_scroll_frame = sf.scrollable_frame

        # Agent Deliverables Auto-Accept Banner (packed dynamically in _refresh_stats)
        self.deliverables_frame = tk.Frame(self._stats_scroll_frame, bg=DARK["card"],
                                           highlightthickness=2, highlightbackground=DARK["success"])
        tk.Label(self.deliverables_frame, text="🎉 Agent Deliverables Ready to Accept!",
                 font=(self._family, 10, "bold"), bg=DARK["card"], fg=DARK["success"]
                 ).pack(anchor=tk.W, padx=10, pady=(10, 2))
        self.deliverables_lbl = tk.Label(self.deliverables_frame, text="0 files completed by your AI agent.",
                                         font=(self._family, 8), bg=DARK["card"], fg=DARK["text_dim"])
        self.deliverables_lbl.pack(anchor=tk.W, padx=10, pady=2)
        self.deliverables_btn = tk.Button(self.deliverables_frame, text="⚡ Auto-Accept & Promote to Done",
                                          font=(self._family, 8, "bold"),
                                          bg=DARK["success"], fg="#07101E", relief="flat", bd=0,
                                          padx=15, pady=6, cursor="hand2",
                                          command=self._bulk_promote_deliverables)
        self.deliverables_btn.pack(fill=tk.X, padx=10, pady=(4, 10))

        # Progress gauge
        self.gauge_card = tk.Frame(self._stats_scroll_frame, bg=DARK["card"],
                                   highlightthickness=1, highlightbackground=DARK["border"])
        self.gauge_card.pack(fill=tk.X, pady=(0, 6), padx=4)
        tk.Label(self.gauge_card, text="🚀  Sprint Release Progress",
                 font=(self._family, 10, "bold"), bg=DARK["card"], fg=DARK["text"]
                 ).pack(anchor=tk.W, padx=10, pady=(10, 2))
        self.gauge_pct_lbl = tk.Label(self.gauge_card, text="0% SHIPPED",
                                      font=(self._family, 18, "bold"),
                                      bg=DARK["card"], fg=DARK["success"])
        self.gauge_pct_lbl.pack(anchor=tk.W, padx=10)
        self.gauge_canvas = tk.Canvas(self.gauge_card, height=18,
                                      bg=DARK["card"], highlightthickness=0)
        self.gauge_canvas.pack(fill=tk.X, padx=10, pady=(4, 6))
        self.gauge_sub_lbl = tk.Label(self.gauge_card, text="0 of 0 components shipped",
                                      font=(self._family, 8), bg=DARK["card"], fg=DARK["text_dim"])
        self.gauge_sub_lbl.pack(anchor=tk.W, padx=10, pady=(0, 10))

        # Git Sync Toggle Button
        self.btn_toggle_git = tk.Button(self._stats_scroll_frame, text="🔄 Show Git & Auto-Promote Tools...", font=(self._family, 8, "bold"),
                                        bg=DARK["surface"], fg=DARK["text_dim"], activebackground=DARK["border"],
                                        activeforeground=DARK["text"], relief="flat", bd=0, pady=5, cursor="hand2",
                                        command=self._toggle_git_sync_card)
        self.btn_toggle_git.pack(fill=tk.X, pady=(0, 6), padx=4)

        # Git Sync & Automation Card (not packed by default!)
        self.git_sync_card = tk.Frame(self._stats_scroll_frame, bg=DARK["card"],
                                      highlightthickness=1, highlightbackground=DARK["border"])
        
        tk.Label(self.git_sync_card, text="🔄  Git Integration & Automation",
                 font=(self._family, 10, "bold"), bg=DARK["card"], fg=DARK["accent"]
                 ).pack(anchor=tk.W, padx=10, pady=(10, 2))
                 
        self.git_status_lbl = tk.Label(self.git_sync_card, text="No active project root linked.",
                                       font=(self._family, 8), bg=DARK["card"], fg=DARK["text_dim"])
        self.git_status_lbl.pack(anchor=tk.W, padx=10, pady=2)
        
        btn_f = tk.Frame(self.git_sync_card, bg=DARK["card"])
        btn_f.pack(fill=tk.X, padx=10, pady=(6, 10))
        
        ttk.Button(btn_f, text="🔄 Refresh Git Status", style="Dk.TButton",
                   command=self._manual_git_refresh).pack(side=tk.LEFT, padx=(0, 4))
                   
        self.auto_sync_btn = tk.Button(btn_f, text="⚡ Auto-Promote Changes",
                                       font=(self._family, 8, "bold"),
                                       bg=DARK["accent"], fg="#07101E", relief="flat", bd=0,
                                       padx=10, pady=5, cursor="hand2",
                                       command=self._bulk_promote_git_changes)
        self.auto_sync_btn.pack(side=tk.LEFT, padx=4)

        self._legend_card = tk.Frame(self._stats_scroll_frame, bg=DARK["card"],
                                     highlightthickness=1, highlightbackground=DARK["border"])
        self._legend_card.pack(fill=tk.X, pady=(0, 6), padx=4)
        self._rebuild_legend_rows()

        self._build_stats_card(self._stats_scroll_frame)

        # Shortcuts Toggle Button
        self.btn_toggle_shortcuts = tk.Button(self._stats_scroll_frame, text="⌨️ Show Keyboard Shortcuts...", font=(self._family, 8, "bold"),
                                              bg=DARK["surface"], fg=DARK["text_dim"], activebackground=DARK["border"],
                                              activeforeground=DARK["text"], relief="flat", bd=0, pady=5, cursor="hand2",
                                              command=self._toggle_shortcuts_card)
        self.btn_toggle_shortcuts.pack(fill=tk.X, pady=(0, 6), padx=4)

        self._build_help_card(self._stats_scroll_frame)

    def _rebuild_legend_rows(self):
        card = self._legend_card
        for w in card.winfo_children():
            w.destroy()

        # Header
        hdr = tk.Frame(card, bg=DARK["card"])
        hdr.pack(fill=tk.X, padx=10, pady=(10, 4))
        tk.Label(hdr, text="🎨  Status Legend & Workflow Guide",
                 font=(self._family, 11, "bold"), bg=DARK["card"], fg=DARK["text"]
                 ).pack(side=tk.LEFT)
        tk.Label(hdr, text="Click swatch to customize",
                 font=(self._family, 7), bg=DARK["card"], fg=DARK["text_dim"]
                 ).pack(side=tk.RIGHT)
        tk.Frame(card, bg=DARK["border"], height=1).pack(fill=tk.X, padx=10)

        # Column headers
        ch = tk.Frame(card, bg=DARK["card"])
        ch.pack(fill=tk.X, padx=10, pady=(4, 0))
        for txt, w in [("Color", 22), ("Icon", 4), ("Status", 14), ("Meaning", 18), ("What to do", 0)]:
            tk.Label(ch, text=txt, font=(self._family, 7, "bold"),
                     bg=DARK["card"], fg=DARK["text_dim"], width=w, anchor=tk.W
                     ).pack(side=tk.LEFT, padx=2)

        tk.Frame(card, bg=DARK["border"], height=1).pack(fill=tk.X, padx=10, pady=(2, 0))

        inner = tk.Frame(card, bg=DARK["card"])
        inner.pack(fill=tk.X, padx=10, pady=(4, 10))

        for status in STATUS_ORDER:
            meta  = STATUS_META[status]
            bg    = self.status_colors.get(status, "#555")
            icon  = STATUS_ICONS[status]
            short_label, meaning, action = STATUS_WORKFLOW[status]

            r2, g2, b2 = int(bg[1:3], 16), int(bg[3:5], 16), int(bg[5:7], 16)
            lum = 0.2126*r2 + 0.7152*g2 + 0.0722*b2
            row_fg = "#0A0A0A" if lum > 80 else "#1A1A1A"

            row = tk.Frame(inner, bg=DARK["card"])
            row.pack(fill=tk.X, pady=3)

            # Clickable color swatch
            sw = tk.Canvas(row, width=22, height=22, bg=DARK["card"], highlightthickness=1,
                           highlightbackground=DARK["border"], cursor="hand2")
            sw.pack(side=tk.LEFT, padx=(0, 4))
            sw.create_rectangle(1, 1, 21, 21, fill=bg, outline="")
            sw.create_text(11, 11, text="🎨", font=("Segoe UI Emoji", 8))
            sw.bind("<Button-1>", lambda e, s=status: self._pick_color(s))

            # Icon badge
            icon_frame = tk.Frame(row, bg=bg, width=28, height=22,
                                  highlightthickness=0)
            icon_frame.pack(side=tk.LEFT, padx=(0, 6))
            icon_frame.pack_propagate(False)
            tk.Label(icon_frame, text=icon, font=("Segoe UI Emoji", 11),
                     bg=bg, fg=row_fg).pack(expand=True)

            # Status name
            tk.Label(row, text=status, font=(self._family, 8, "bold"),
                     bg=DARK["card"], fg=DARK["text"], width=12, anchor=tk.W
                     ).pack(side=tk.LEFT, padx=(0, 4))

            # Meaning
            tk.Label(row, text=meaning, font=(self._family, 7),
                     bg=DARK["card"], fg=DARK["text_dim"], width=14, anchor=tk.W
                     ).pack(side=tk.LEFT, padx=(0, 4))

            # Action
            tk.Label(row, text=action, font=(self._family, 7),
                     bg=DARK["card"], fg=DARK["text_dim"], anchor=tk.W, wraplength=130
                     ).pack(side=tk.LEFT, fill=tk.X, expand=True)

            # Click row to filter
            row.bind("<Button-1>", lambda e, s=status: self._set_filter(s))

    def _pick_color(self, status: str):
        current = self.status_colors.get(status, "#888888")
        result  = colorchooser.askcolor(color=current, title=f"Color for {status}",
                                        parent=self.root)
        if result and result[1]:
            self._custom_colors[status]   = result[1]
            self.status_colors[status]    = result[1]
            self._persist_config()
            self._reconfigure_tree_tags()
            self._rebuild_filter_pills()
            self._rebuild_legend_rows()
            self._repopulate()
            self.sb_left.config(text=f"🎨 Custom color saved for {status}")

    def _build_stats_card(self, parent):
        card = tk.Frame(parent, bg=DARK["card"], highlightthickness=1,
                        highlightbackground=DARK["border"])
        card.pack(fill=tk.X, pady=(0, 6), padx=4)
        tk.Label(card, text="📊  Project Statistics", font=(self._family, 11, "bold"),
                 bg=DARK["card"], fg=DARK["text"]).pack(anchor=tk.W, padx=10, pady=(10, 6))
        tk.Frame(card, bg=DARK["border"], height=1).pack(fill=tk.X, padx=10)
        self.stats_inner = tk.Frame(card, bg=DARK["card"])
        self.stats_inner.pack(fill=tk.X, padx=10, pady=(6, 10))

    def _build_help_card(self, parent):
        self.shortcuts_card = tk.Frame(parent, bg=DARK["card"], highlightthickness=1,
                                       highlightbackground=DARK["border"])
        card = self.shortcuts_card
        tk.Label(card, text="⌨️  Shortcuts", font=(self._family, 11, "bold"),
                 bg=DARK["card"], fg=DARK["text"]).pack(anchor=tk.W, padx=10, pady=(10, 6))
        tk.Frame(card, bg=DARK["border"], height=1).pack(fill=tk.X, padx=10)
        shortcuts = [
            ("Ctrl+O",       "Open workbook"),
            ("Ctrl+S",       "Save changes"),
            ("Ctrl+F",       "Focus search bar"),
            ("Ctrl+±",       "Zoom in / out"),
            ("1–9 / C N S A","Quick status key shortcuts"),
            ("Right-click",  "Context menu with all actions"),
            ("Dbl-click Status col", "Quick status picker popup"),
            ("🎨 Theme btn",  "Cycle through 5 color palettes"),
        ]
        inner = tk.Frame(card, bg=DARK["card"])
        inner.pack(fill=tk.X, padx=10, pady=(6, 10))
        for key, desc in shortcuts:
            r = tk.Frame(inner, bg=DARK["card"])
            r.pack(fill=tk.X, pady=1)
            tk.Label(r, text=key, font=(self._mono, 7, "bold"),
                     bg=DARK["card"], fg=DARK["accent"], width=22, anchor=tk.W).pack(side=tk.LEFT)
            tk.Label(r, text=desc, font=(self._family, 7),
                     bg=DARK["card"], fg=DARK["text_dim"]).pack(side=tk.LEFT)

    # ─────────────────────────────────────────────────────────────────
    # Status bar
    # ─────────────────────────────────────────────────────────────────
    def _build_statusbar(self):
        bar = tk.Frame(self.root, bg=DARK["surface"], height=26)
        bar.pack(fill=tk.X, side=tk.BOTTOM)
        bar.pack_propagate(False)
        tk.Frame(bar, bg=DARK["border"], height=1).pack(fill=tk.X, side=tk.TOP)
        self.sb_left  = tk.Label(bar, text="Ready", font=(self._family, 8),
                                  bg=DARK["surface"], fg=DARK["text_dim"], anchor=tk.W)
        self.sb_left.pack(side=tk.LEFT, padx=10)
        self.sb_right = tk.Label(bar, text="", font=(self._family, 8),
                                  bg=DARK["surface"], fg=DARK["accent"], anchor=tk.E)
        self.sb_right.pack(side=tk.RIGHT, padx=10)

    # ─────────────────────────────────────────────────────────────────
    # Rotating tips
    # ─────────────────────────────────────────────────────────────────
    def _rotate_tips(self):
        if hasattr(self, "tip_lbl"):
            self.tip_lbl.config(text=self.tips[self.current_tip_idx])
            self.current_tip_idx = (self.current_tip_idx + 1) % len(self.tips)
        self.root.after(14000, self._rotate_tips)

    # ═══════════════════════════════════════════════════════════════════
    # Excel I/O
    # ═══════════════════════════════════════════════════════════════════
    def _open_file(self):
        p = filedialog.askopenfilename(
            title="Open Project Skeleton (.xlsx)",
            filetypes=[("Excel Workspace", "*.xlsx"), ("All Files", "*.*")])
        if p:
            self._load(p)

    def _open_directory(self):
        p = filedialog.askdirectory(title="Select Project Repository Folder")
        if p:
            self._load(p)

    def _load(self, path: str):
        path_obj = Path(path).resolve()
        
        if path_obj.is_dir() or path_obj.suffix.lower() != ".xlsx":
            # Repository mode!
            self.repository_mode = True
            if path_obj.is_file():
                # Search upwards for .git or 90_AUDITME
                root_candidate = path_obj.parent
                while root_candidate != root_candidate.parent:
                    if (root_candidate / "90_AUDITME").is_dir() or (root_candidate / ".git").is_dir():
                        break
                    root_candidate = root_candidate.parent
                self.linked_project_root = str(root_candidate)
            else:
                self.linked_project_root = str(path_obj)
            
            self.xlsx_path = None
            self.dirty = False
            self.current_editor_idx = None
            self.sb_right.config(text="")
            self.editor_fields_frame.pack_forget()
            self.editor_placeholder.pack(fill=tk.BOTH, expand=True)
            
            # Scan files directly
            self._scan_repository()
            
        else:
            # Excel mode
            self.repository_mode = False
            try:
                self.workbook = openpyxl.load_workbook(path)
            except Exception as exc:
                messagebox.showerror("Error", f"Cannot open workbook:\n{exc}")
                return

            self.xlsx_path = path
            self.dirty = False
            self.current_editor_idx = None
            self.sb_right.config(text="")
            self.editor_fields_frame.pack_forget()
            self.editor_placeholder.pack(fill=tk.BOTH, expand=True)

            if "Control Map" not in self.workbook.sheetnames:
                messagebox.showwarning("Warning", "No 'Control Map' sheet found.")
                return

            self._parse()

            # Resolve linked project root from config or try auto-link
            project_roots = self._config.setdefault("project_roots", {})
            abs_xlsx = str(Path(path).resolve())
            if abs_xlsx in project_roots:
                self.linked_project_root = project_roots[abs_xlsx]
            else:
                self.linked_project_root = None
                if "DreamTeamHQ" in Path(path).name:
                    auto_path = "C:\\Projects\\DreamTeamHQ"
                    if os.path.isdir(auto_path):
                        self.linked_project_root = auto_path
                        project_roots[abs_xlsx] = auto_path
                        self._persist_config()
                if not self.linked_project_root:
                    self.linked_project_root = str(Path(path).parent.resolve())

        self._refresh_git_status()
        self._repopulate()
        self._refresh_stats()
        self._update_link_label()
        self._update_git_sync_card()

        if getattr(self, "repository_mode", False):
            name = Path(self.linked_project_root).name
            n = len(self.rows)
            self.root.title(f"AuditME Desktop (Native Repo) — {name}")
            self.subtitle_lbl.config(text=f"  {name} (Native Repo)  •  {n} entries")
            self.sb_left.config(text=f"Indexed Repository: {name}  —  {n} markdown tasks")
        else:
            name = Path(path).name
            n    = len(self.rows)
            self.root.title(f"AuditME Desktop — {name}")
            self.subtitle_lbl.config(text=f"  {name}  •  {n} entries")
            self.sb_left.config(text=f"Loaded: {name}  —  {n} rows")

    def _scan_repository(self):
        self.rows = []
        project_root = Path(self.linked_project_root)
        if not project_root.exists():
            return
            
        ignored_dirs = {".git", ".venv", "node_modules", ".tmp.driveupload", "90_AUDITME", ".antigravity", "Project_Startup", "__pycache__"}
        
        md_files = []
        def _recurse(dir_path):
            try:
                for item in dir_path.iterdir():
                    if item.is_dir():
                        if item.name not in ignored_dirs and not item.name.startswith(".tmp-"):
                            _recurse(item)
                    elif item.is_file() and item.suffix.lower() == ".md":
                        md_files.append(item)
            except Exception:
                pass
                
        _recurse(project_root)
        md_files.sort()
        
        for ri, file_path in enumerate(md_files, 2):
            rel_path = file_path.relative_to(project_root).as_posix()
            try:
                content = file_path.read_text(encoding="utf-8", errors="replace")
                clean_content = content.replace("\uFEFF", "").strip()
                if clean_content.startswith("---"):
                    parts = clean_content.split("---", 2)
                    if len(parts) >= 3:
                        frontmatter_str = parts[1]
                        body = parts[2]
                        
                        frontmatter_data = {}
                        for line in frontmatter_str.splitlines():
                            if ":" in line:
                                key, val = line.split(":", 1)
                                frontmatter_data[key.strip()] = val.strip()
                        
                        status = frontmatter_data.get("STATUS")
                        if status:
                            purpose = ""
                            for line in body.splitlines():
                                if line.strip().startswith("#"):
                                    purpose = line.replace("#", "").strip()
                                    break
                            
                            d = {
                                "_ri": ri,
                                "_abs_path": str(file_path),
                                "Path": rel_path,
                                "Type": "file",
                                "Status": status,
                                "Color": STATUS_META.get(status, {}).get("label", "Gray"),
                                "Lock State": frontmatter_data.get("LOCK_STATE", "UNLOCKED"),
                                "Phase": frontmatter_data.get("READY_FOR", ""),
                                "Owner": frontmatter_data.get("OWNER", ""),
                                "Purpose": purpose or frontmatter_data.get("Purpose", ""),
                                "Build Rule": frontmatter_data.get("RULE", ""),
                                "Notes": frontmatter_data.get("LAST_REVIEWED", "")
                            }
                            self.rows.append(d)
            except Exception:
                pass

    def _parse(self):
        ws = self.workbook["Control Map"]
        hdrs: list[str] = []
        self.rows = []
        for ri, row in enumerate(ws.iter_rows(min_row=1, values_only=False), 1):
            if ri == 1:
                hdrs = [(c.value or f"_col{i}") for i, c in enumerate(row)]
                continue
            d = {"_ri": ri}
            for i, cell in enumerate(row):
                key = hdrs[i] if i < len(hdrs) else f"_col{i}"
                val = cell.value or ""
                if key == "Path" and isinstance(val, str):
                    val = val.strip().rstrip("/")
                d[key] = val
            self.rows.append(d)

    def _save(self):
        if getattr(self, "repository_mode", False):
            errors = []
            for d in self.rows:
                abs_path_str = d.get("_abs_path")
                if not abs_path_str:
                    continue
                abs_path = Path(abs_path_str)
                try:
                    if not abs_path.exists():
                        # Create file with frontmatter template
                        abs_path.parent.mkdir(parents=True, exist_ok=True)
                        frontmatter_lines = [
                            "---",
                            f"STATUS: {d.get('Status', 'PLACEHOLDER')}",
                            f"OWNER: {d.get('Owner', '')}",
                            f"READY_FOR: {d.get('Phase', '')}",
                            f"LAST_REVIEWED: {d.get('Notes', '')}",
                            f"LOCK_STATE: {d.get('Lock State', 'UNLOCKED')}",
                            "---",
                            "",
                            f"# {d.get('Purpose') or Path(abs_path).stem or 'Untitled File'}",
                            ""
                        ]
                        abs_path.write_text("\n".join(frontmatter_lines), encoding="utf-8")
                        continue

                    content = abs_path.read_text(encoding="utf-8", errors="replace")
                    clean_content = content.replace('\uFEFF', '')
                    if not clean_content.startswith("---"):
                        frontmatter_lines = [
                            "---",
                            f"STATUS: {d.get('Status', 'PLACEHOLDER')}",
                            f"OWNER: {d.get('Owner', '')}",
                            f"READY_FOR: {d.get('Phase', '')}",
                            f"LAST_REVIEWED: {d.get('Notes', '')}",
                            f"LOCK_STATE: {d.get('Lock State', 'UNLOCKED')}",
                            "---"
                        ]
                        new_content = "\n".join(frontmatter_lines) + "\n\n" + content
                        abs_path.write_text(new_content, encoding="utf-8")
                        continue
                        
                    parts = clean_content.split("---", 2)
                    frontmatter_str = parts[1]
                    body = parts[2]
                    
                    frontmatter_data = {}
                    frontmatter_order = []
                    for line in frontmatter_str.splitlines():
                        if ":" in line:
                            key, val = line.split(":", 1)
                            k = key.strip()
                            frontmatter_data[k] = val.strip()
                            frontmatter_order.append(k)
                        else:
                            frontmatter_order.append(line)
                            
                    # Update fields
                    frontmatter_data["STATUS"] = d.get("Status", "PLACEHOLDER")
                    frontmatter_data["OWNER"] = d.get("Owner", "")
                    frontmatter_data["READY_FOR"] = d.get("Phase", "")
                    frontmatter_data["LOCK_STATE"] = d.get("Lock State", "UNLOCKED")
                    frontmatter_data["LAST_REVIEWED"] = d.get("Notes", "")
                    if d.get("Build Rule"):
                        frontmatter_data["RULE"] = d.get("Build Rule")
                        if "RULE" not in frontmatter_order:
                            frontmatter_order.append("RULE")
                            
                    for k in ["STATUS", "OWNER", "READY_FOR", "LOCK_STATE", "LAST_REVIEWED"]:
                        if k not in frontmatter_order:
                            frontmatter_order.append(k)
                            
                    # Reconstruct frontmatter
                    new_lines = []
                    for item in frontmatter_order:
                        if item in frontmatter_data:
                            new_lines.append(f"{item}: {frontmatter_data[item]}")
                        else:
                            new_lines.append(item)
                            
                    new_frontmatter_str = "\n".join(new_lines)
                    new_content = "---\n" + new_frontmatter_str + "\n---\n" + body
                    abs_path.write_text(new_content, encoding="utf-8")
                except Exception as exc:
                    errors.append(f"{d.get('Path')}: {exc}")
                    
            if errors:
                messagebox.showerror("Save Errors", "Errors occurred while saving some files:\n" + "\n".join(errors))
            else:
                self.dirty = False
                self.sb_right.config(text="✅ Saved directly to Repository Markdown files!", fg=DARK["success"])
                self.root.after(3000, lambda: self.sb_right.config(text="", fg=DARK["accent"]))
                self._repopulate()
        else:
            if not self.workbook or not self.xlsx_path:
                messagebox.showinfo("Nothing to save", "No workbook loaded.")
                return
            ws   = self.workbook["Control Map"]
            hdrs = [cell.value for cell in ws[1]]
            ws.delete_rows(2, ws.max_row)
            for index, d in enumerate(self.rows, 2):
                d["_ri"] = index
                status   = d.get("Status", "")
                meta     = STATUS_META.get(status)
                for col_idx, hdr in enumerate(hdrs, 1):
                    if hdr in d:
                        val = d[hdr]
                        if hdr == "Path" and d.get("Type", "").lower() in ("folder", "repo") and isinstance(val, str):
                            if not val.endswith("/"):
                                        val += "/"
                        ws.cell(row=index, column=col_idx).value = val
                if meta:
                    bg_hex = self.status_colors.get(status, "#FFFFFF").lstrip("#")
                    if "Color" in hdrs:
                        ws.cell(row=index, column=hdrs.index("Color")+1).value = meta["label"]
                    if "Lock State" in hdrs:
                        ws.cell(row=index, column=hdrs.index("Lock State")+1).value = meta["meaning"]
                    fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
                    for c in range(1, len(hdrs)+1):
                        ws.cell(row=index, column=c).fill = fill
            try:
                self.workbook.save(self.xlsx_path)
                self.dirty = False
                self.sb_right.config(text="✅ Saved!", fg=DARK["success"])
                self.root.after(3000, lambda: self.sb_right.config(text="", fg=DARK["accent"]))
                self._repopulate()
            except Exception as exc:
                messagebox.showerror("Save Error", str(exc))

    # ═══════════════════════════════════════════════════════════════════
    # IDE Tree Renderer
    # ═══════════════════════════════════════════════════════════════════
    def _repopulate(self):
        for ch in self.tree.get_children():
            self.tree.delete(ch)
        if not self.rows:
            return

        query = self.search_var.get().lower().strip()
        filt  = self.active_filter
        focus = self.focus_var.get()
        project_root = Path(self.linked_project_root) if self.linked_project_root else (Path(self.xlsx_path).parent if self.xlsx_path else Path("."))
        project_name = Path(self.linked_project_root).name if getattr(self, "repository_mode", False) else (Path(self.xlsx_path).stem if self.xlsx_path else "workspace")

        def _passes(d):
            status  = d.get("Status", "")
            path    = d.get("Path",    "").lower()
            purpose = d.get("Purpose","").lower()
            if filt and filt != "ALL" and status != filt:
                return False
            if focus and status not in ("CONSTRUCTION","NEXT","HUMAN_GATE","PLACEHOLDER"):
                return False
            if query and query not in path and query not in purpose:
                return False
            return True

        # Ancestors of matching rows
        passing: set[str] = set()
        for d in self.rows:
            if _passes(d):
                p = d.get("Path","").strip().rstrip("/")
                if p:
                    parts = p.split("/")
                    for i in range(1, len(parts)+1):
                        passing.add("/".join(parts[:i]))

        if not passing and (query or filt or focus):
            self.tree_count_lbl.config(text="0 matches")
            return

        # Folder hot-map
        folder_hot: dict[str, str] = {}
        for d in self.rows:
            s = d.get("Status","")
            if s in ("CONSTRUCTION","NEXT","PLACEHOLDER"):
                p     = d.get("Path","").strip().rstrip("/")
                parts = p.split("/")
                for i in range(1, len(parts)):
                    anc = "/".join(parts[:i])
                    if anc not in folder_hot:
                        folder_hot[anc] = s

        # Virtual parents
        explicit_paths = {d.get("Path","").strip().rstrip("/") for d in self.rows if d.get("Path")}
        virtual_paths: set[str] = set()
        for d in self.rows:
            p = d.get("Path","").strip().rstrip("/")
            for i in range(1, len(p.split("/"))):
                pp = "/".join(p.split("/")[:i])
                if pp not in explicit_paths:
                    virtual_paths.add(pp)

        # Build render list
        render: list[dict] = []
        for i, d in enumerate(self.rows):
            dc = d.copy()
            dc["_original_idx"] = i
            dc["_is_virtual"]   = False
            render.append(dc)
        for vp in virtual_paths:
            render.append({
                "Path": vp, "Type": "folder",
                "Status": folder_hot.get(vp,"FROZEN"),
                "_is_virtual": True, "_original_idx": None,
                "Purpose":"","Owner":"","Phase":"",
            })
        render.sort(key=lambda x: (x.get("Path","").count("/"), x.get("Path","").lower()))

        # Root node
        root_node = self.tree.insert("", tk.END, iid="__root__",
            text=f"📦  {project_name}/",
            values=("  workspace",), tags=("workspace_root",), open=True)

        path_to_iid: dict[str, str] = {}
        count = 0

        for item in render:
            path      = item.get("Path","").strip().rstrip("/")
            status    = item.get("Status","")
            is_virt   = item.get("_is_virtual", False)
            etype     = item.get("Type","").lower()

            if passing and path not in passing:
                continue
            if is_virt and (query or filt or focus) and path not in passing:
                continue

            parts      = path.split("/")
            name       = parts[-1]
            parent_iid = "__root__"
            if len(parts) > 1:
                parent_iid = path_to_iid.get("/".join(parts[:-1]), "__root__")

            # ── Disk check & size check ──
            exists = False
            size_str = ""
            if not is_virt and etype not in ("folder","repo") and (self.xlsx_path or getattr(self, "repository_mode", False)):
                fp = project_root / path
                exists = fp.is_file()
                if exists:
                    try:
                        sz = fp.stat().st_size
                        if sz == 0:
                            sz_formatted = "Empty"
                        elif sz < 1024:
                            sz_formatted = f"{sz} B"
                        elif sz < 1024 * 1024:
                            sz_formatted = f"{sz/1024:.1f} KB"
                        else:
                            sz_formatted = f"{sz/(1024*1024):.1f} MB"
                        # Only show ▲/▼ when there is a real measured change vs Git baseline.
                        # No stub guessing — the status column on the right already shows STUB.
                        growth_indicator = ""
                        orig_sz = self._get_cached_original_size(path)
                        if orig_sz is not None:
                            diff = sz - orig_sz
                            if diff > 50:
                                diff_str = f"+{diff} B" if diff < 1024 else f"+{diff/1024:.1f} KB"
                                growth_indicator = f"  ▲{diff_str}"
                            elif diff < -50:
                                abs_diff = abs(diff)
                                diff_str = f"-{abs_diff} B" if abs_diff < 1024 else f"-{abs_diff/1024:.1f} KB"
                                growth_indicator = f"  ▼{diff_str}"
                        size_str = f" ({sz_formatted}{growth_indicator})"
                    except Exception:
                        size_str = " (Error)"

            # ── File icon ──
            if etype in ("folder","repo") or is_virt:
                file_icon = "📁"
            elif etype == "test":
                file_icon = "🧪"
            elif name.endswith(".md"):
                file_icon = "📝"
            elif name.endswith((".py",".js",".ts",".jsx",".tsx",".go",".rs")):
                file_icon = "📄"
            elif name.endswith((".json",".yaml",".yml",".toml",".env")):
                file_icon = "⚙️"
            elif name.endswith((".css",".scss")):
                file_icon = "🎨"
            else:
                file_icon = "📄"

            # ── Status icon (the workflow icon, NOT a circle) ──
            eff_status   = folder_hot.get(path, status) if etype in ("folder","repo") else status
            status_icon  = STATUS_ICONS.get(eff_status, "?")

            # ── Disk indicator — clearly text, not a circle ──
            git_state = self.git_changed_files.get(path)  # 'modified' or 'untracked'
            not_in_main = path in self.branch_diff_files
            git_badge = ""
            if git_state == "modified":
                git_badge = " 🔄"
            elif git_state == "untracked":
                git_badge = " ➕"
            
            branch_badge = " 🌿" if not_in_main else ""

            if etype not in ("folder","repo") and not is_virt:
                disk_str = f"  💾{size_str}{git_badge}{branch_badge}" if exists else f"  ✗{git_badge}{branch_badge}"
            else:
                disk_str = f"{git_badge}{branch_badge}"

            # ── Display text ──
            if etype in ("folder","repo") or is_virt:
                display_text = f"{file_icon}  {name}/"
            else:
                display_text = f"{file_icon}  {name}{disk_str}"

            # ── Status column ──
            if is_virt:
                status_col = "  📁 Folder"
                tags = ("virtual_folder",)
            else:
                short_label = STATUS_WORKFLOW.get(eff_status, ("","",""))[0]
                status_col  = f"{status_icon} {short_label}"
                extra_badges = []
                if git_state:
                    extra_badges.append(git_state.upper())
                if not_in_main:
                    extra_badges.append("NOT IN MAIN")
                if extra_badges:
                    status_col += f" ({', '.join(extra_badges)})"
                tags = (status,) if status in STATUS_META else ()
                if git_state == "modified":
                    tags = tags + ("git_modified",)
                elif git_state == "untracked":
                    tags = tags + ("git_untracked",)
                if not_in_main:
                    tags = tags + ("not_in_main",)

            iid = f"v_{path}" if is_virt else str(item["_original_idx"])

            try:
                node = self.tree.insert(parent_iid, tk.END,
                    iid=iid, text=display_text,
                    values=(status_col,), tags=tags)
            except tk.TclError:
                node = iid

            if etype in ("folder","repo") or is_virt:
                self.tree.item(node, open=True)

            path_to_iid[path] = node
            if not is_virt:
                count += 1

        self.tree_count_lbl.config(text=f"{count} items")

    # ─────────────────────────────────────────────────────────────────
    # Expand / Collapse
    # ─────────────────────────────────────────────────────────────────
    def _expand_all_nodes(self):
        def _ex(item):
            self.tree.item(item, open=True)
            for ch in self.tree.get_children(item):
                _ex(ch)
        for item in self.tree.get_children():
            _ex(item)

    def _collapse_all_nodes(self):
        def _co(item):
            self.tree.item(item, open=False)
            for ch in self.tree.get_children(item):
                _co(ch)
        for item in self.tree.get_children():
            _co(item)

    # ═══════════════════════════════════════════════════════════════════
    # Disk stub / open
    # ═══════════════════════════════════════════════════════════════════
    def _create_file_stub(self, rel_path, purpose, rule, owner, status):
        if not self.xlsx_path and not getattr(self, "repository_mode", False):
            return
        project_root = Path(self.linked_project_root) if self.linked_project_root else (Path(self.xlsx_path).parent if self.xlsx_path else Path("."))
        abs_path = project_root / rel_path
        try:
            abs_path.parent.mkdir(parents=True, exist_ok=True)
            ext = abs_path.suffix.lower()
            if ext in (".py",".sh",".yaml",".yml",".toml",".ini",".md"):
                cs, ce = "# ", ""
            elif ext in (".js",".ts",".jsx",".tsx",".c",".cpp",".h",".java",".cs",".css",".scss",".go"):
                cs, ce = "// ", ""
            elif ext in (".html",".xml",".svg"):
                cs, ce = "<!-- ", " -->"
            else:
                cs, ce = "# ", ""
            content = "\n".join([
                f"{cs}═════════════════════════════════════════════════{ce}",
                f"{cs}File:    {rel_path}{ce}",
                f"{cs}Status:  {status}{ce}",
                f"{cs}Owner:   {owner or 'Unassigned'}{ce}",
                f"{cs}Purpose: {purpose or 'Not specified'}{ce}",
                f"{cs}Rule:    {rule or 'Follow standard practices.'}{ce}",
                f"{cs}═════════════════════════════════════════════════{ce}", "",
            ])
            with open(abs_path, "w", encoding="utf-8") as f:
                f.write(content)
            self.sb_right.config(text="✨ Stub created!", fg=DARK["success"])
            self._repopulate()
            if self.current_editor_idx is not None:
                self._load_detail_panel(self.current_editor_idx)
        except Exception as exc:
            messagebox.showerror("Error", str(exc))

    def _open_file_in_editor(self, rel_path):
        if not self.xlsx_path and not getattr(self, "repository_mode", False):
            return
        project_root = Path(self.linked_project_root) if self.linked_project_root else (Path(self.xlsx_path).parent if self.xlsx_path else Path("."))
        abs_path = project_root / rel_path
        if not abs_path.exists():
            messagebox.showwarning("Missing", f"File not on disk: {rel_path}")
            return
        try:
            os.startfile(abs_path)
        except Exception:
            import subprocess
            subprocess.run(["cmd","/c","start","",str(abs_path)], shell=True)

    def _create_stub_for_selected(self):
        if self.current_editor_idx is not None:
            d = self.rows[self.current_editor_idx]
            self._create_file_stub(d.get("Path",""), d.get("Purpose",""),
                                   d.get("Build Rule",""), d.get("Owner",""), d.get("Status",""))

    def _open_selected_in_editor(self):
        if self.current_editor_idx is not None:
            self._open_file_in_editor(self.rows[self.current_editor_idx].get("Path",""))

    # ═══════════════════════════════════════════════════════════════════
    # Add Node Dialog
    # ═══════════════════════════════════════════════════════════════════
    def _on_add_node_click(self):
        if not getattr(self, "repository_mode", False) and (not self.workbook or not self.xlsx_path):
            messagebox.showinfo("No Workspace","Please open an Excel workbook or Project Repository first.")
            return

        prefill_path = ""
        sel = self.tree.selection()
        if sel:
            iid = sel[0]
            if iid == "__root__":
                prefill_path = ""
            elif iid.startswith("v_"):
                prefill_path = iid[2:] + "/"
            else:
                try:
                    d     = self.rows[int(iid)]
                    p     = d.get("Path","")
                    etype = d.get("Type","").lower()
                    if etype in ("folder","repo"):
                        prefill_path = p + "/"
                    else:
                        parts = p.split("/")
                        prefill_path = "/".join(parts[:-1]) + "/" if len(parts) > 1 else ""
                except Exception:
                    pass

        dlg = tk.Toplevel(self.root)
        dlg.title("➕ Add File/Folder Node")
        dlg.geometry("560x620")
        dlg.configure(bg=DARK["bg"])
        dlg.transient(self.root)
        dlg.grab_set()

        tk.Label(dlg, text="➕ Add New Node to Registry",
                 font=(self._family, 12, "bold"), bg=DARK["bg"], fg=DARK["accent"]
                 ).pack(anchor=tk.W, padx=20, pady=(15, 4))

        # Path guidance box
        guide = tk.Frame(dlg, bg=DARK["card"], highlightthickness=1,
                         highlightbackground=DARK["accent"])
        guide.pack(fill=tk.X, padx=20, pady=(0, 10))
        lines = [
            "💡 Path is RELATIVE to the folder where your .xlsx file lives.",
            "   Use forward slashes /  and no leading slash.",
            "   Example:  src/components/Header.jsx",
        ]
        if prefill_path:
            lines.append(f"   Pre-filled from selection:  {prefill_path}")
        for line in lines:
            fg = DARK["accent"] if "Pre-filled" in line else DARK["text_dim"]
            tk.Label(guide, text=line, font=(self._family, 8), bg=DARK["card"], fg=fg,
                     anchor=tk.W).pack(anchor=tk.W, padx=10, pady=1)
        tk.Label(guide, text="", bg=DARK["card"]).pack()

        grid = tk.Frame(dlg, bg=DARK["bg"])
        grid.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)
        grid.columnconfigure(1, weight=1)

        path_var   = tk.StringVar(value=prefill_path)
        type_var   = tk.StringVar(value="file")
        status_var = tk.StringVar(value="CONSTRUCTION")
        owner_var  = tk.StringVar()
        phase_var  = tk.StringVar()

        rows_defs = [
            ("Relative Path:", path_var, "entry", "e.g. src/components/button.py"),
            ("Node Type:",     type_var, "combo_type", ""),
            ("Initial Status:",status_var,"combo_status",""),
            ("Owner:",         owner_var, "entry", "e.g. James"),
            ("Phase/Sprint:",  phase_var, "entry", "e.g. Sprint 1"),
        ]
        for r, (lbl, var, ftype, ph) in enumerate(rows_defs):
            tk.Label(grid, text=lbl, font=(self._family, 8, "bold"),
                     bg=DARK["bg"], fg=DARK["text_dim"]
                     ).grid(row=r, column=0, sticky="w", pady=6, padx=(0, 10))
            if ftype == "entry":
                ent = tk.Entry(grid, textvariable=var, font=self.f,
                               bg=DARK["surface"], fg=DARK["text"],
                               insertbackground=DARK["text"], relief="flat",
                               highlightthickness=1, highlightcolor=DARK["accent"],
                               highlightbackground=DARK["border"])
                ent.grid(row=r, column=1, sticky="ew", pady=6)
                if ph:
                    tk.Label(grid, text=ph, font=(self._family, 7),
                             bg=DARK["bg"], fg=DARK["text_dim"]
                             ).grid(row=r, column=1, sticky="e", padx=5)
            elif ftype == "combo_type":
                ttk.Combobox(grid, textvariable=var, values=["file","folder","test"],
                             state="readonly", font=self.f).grid(row=r, column=1, sticky="ew", pady=6)
            elif ftype == "combo_status":
                ttk.Combobox(grid, textvariable=var, values=STATUS_ORDER,
                             state="readonly", font=self.f).grid(row=r, column=1, sticky="ew", pady=6)

        for r, (lbl, h) in enumerate([("Purpose:", 3), ("Build Rule:", 3)], start=5):
            tk.Label(grid, text=lbl, font=(self._family, 8, "bold"),
                     bg=DARK["bg"], fg=DARK["text_dim"]).grid(row=r, column=0, sticky="nw", pady=6)
        txt_purpose = tk.Text(grid, height=3, font=self.f, bg=DARK["surface"], fg=DARK["text"],
                              insertbackground=DARK["text"], relief="flat", highlightthickness=1,
                              highlightcolor=DARK["accent"], highlightbackground=DARK["border"], wrap=tk.WORD)
        txt_purpose.grid(row=5, column=1, sticky="ew", pady=6)
        txt_rule = tk.Text(grid, height=3, font=self.f, bg=DARK["surface"], fg=DARK["text"],
                           insertbackground=DARK["text"], relief="flat", highlightthickness=1,
                           highlightcolor=DARK["accent"], highlightbackground=DARK["border"], wrap=tk.WORD)
        txt_rule.grid(row=6, column=1, sticky="ew", pady=6)

        def add_node():
            pv = path_var.get().strip().rstrip("/").replace("\\","/")
            if not pv:
                messagebox.showwarning("Required", "Please enter a relative path.")
                return
            if any(r.get("Path","").strip().rstrip("/") == pv for r in self.rows):
                messagebox.showerror("Duplicate", f"Path already exists:\n{pv}")
                return
            meta = STATUS_META.get(status_var.get(), {})
            node = {
                "Path": pv, "Type": type_var.get(), "Status": status_var.get(),
                "Color": meta.get("label",""), "Lock State": meta.get("meaning",""),
                "Phase": phase_var.get().strip(), "Owner": owner_var.get().strip(),
                "Purpose": txt_purpose.get("1.0","end-1c").strip(),
                "Build Rule": txt_rule.get("1.0","end-1c").strip(),
                "Notes": "", "_ri": len(self.rows)+2,
            }
            if getattr(self, "repository_mode", False):
                node["_abs_path"] = str(Path(self.linked_project_root) / pv)
            self.rows.append(node)
            self.dirty = True
            self.sb_right.config(text="● Unsaved changes", fg=DARK["danger"])
            self._repopulate()
            self._refresh_stats()
            new_iid = str(len(self.rows)-1)
            try:
                self.tree.selection_set(new_iid)
                self.tree.focus(new_iid)
                self.tree.see(new_iid)
            except Exception:
                pass
            dlg.destroy()

        bf = tk.Frame(dlg, bg=DARK["bg"])
        bf.pack(fill=tk.X, padx=20, pady=15)
        tk.Button(bf, text="✨ Add to Registry", font=(self._family, 9, "bold"),
                  bg=DARK["accent"], fg="#07101E", relief="flat", bd=0, padx=20, pady=8,
                  cursor="hand2", command=add_node).pack(side=tk.LEFT)
        tk.Button(bf, text="Cancel", font=(self._family, 9),
                  bg=DARK["card"], fg=DARK["text"], relief="flat", bd=1, padx=20, pady=8,
                  cursor="hand2", command=dlg.destroy).pack(side=tk.RIGHT)

    # ═══════════════════════════════════════════════════════════════════
    # Gauge helper
    # ═══════════════════════════════════════════════════════════════════
    def _draw_rounded_rect(self, canvas, x1, y1, x2, y2, radius=8, **kw):
        pts = [x1+radius,y1,x1+radius,y1,x2-radius,y1,x2-radius,y1,x2,y1,x2,y1+radius,
               x2,y1+radius,x2,y2-radius,x2,y2-radius,x2,y2,x2-radius,y2,x2-radius,y2,
               x1+radius,y2,x1+radius,y2,x1,y2,x1,y2-radius,x1,y2-radius,x1,y1+radius,
               x1,y1+radius,x1,y1]
        return canvas.create_polygon(pts, **kw, smooth=True)

    # ═══════════════════════════════════════════════════════════════════
    # Stats refresh
    # ═══════════════════════════════════════════════════════════════════
    def _refresh_stats(self):
        if not self.rows:
            return

        # Check for newly completed agent deliverables (files in CONSTRUCTION/NEXT/PLACEHOLDER that now physically exist on disk and have actual content > 400 bytes)
        ready_count = 0
        if self.linked_project_root and os.path.isdir(self.linked_project_root):
            proj_root = Path(self.linked_project_root)
            for d in self.rows:
                path = d.get("Path", "")
                status = d.get("Status", "")
                if status in ("CONSTRUCTION", "NEXT", "PLACEHOLDER") and d.get("Type", "").lower() not in ("folder", "repo"):
                    abs_path = proj_root / path
                    if abs_path.is_file() and abs_path.stat().st_size > 400:
                        ready_count += 1
                        
        if ready_count > 0 and hasattr(self, "deliverables_frame"):
            self.deliverables_lbl.config(text=f"Found {ready_count} completed components with new code!")
            self.gauge_card.pack_forget()
            if hasattr(self, "btn_toggle_git"):
                self.btn_toggle_git.pack_forget()
            self.git_sync_card.pack_forget()
            
            self.deliverables_frame.pack(fill=tk.X, pady=(0, 6), padx=4)
            self.gauge_card.pack(fill=tk.X, pady=(0, 6), padx=4)
            if hasattr(self, "btn_toggle_git"):
                self.btn_toggle_git.pack(fill=tk.X, pady=(0, 6), padx=4)
            self._update_git_sync_card_packing()
        elif hasattr(self, "deliverables_frame"):
            self.deliverables_frame.pack_forget()

        shipped = sum(1 for d in self.rows if d.get("Status") in ("SHIP_READY","FROZEN","ACTIVE_V1"))
        total   = len(self.rows)
        pct     = shipped / total * 100 if total else 0

        self.gauge_pct_lbl.config(text=f"{pct:.0f}% SHIPPED")
        self.gauge_sub_lbl.config(text=f"{shipped} of {total} components shipped")
        self.gauge_canvas.delete("all")
        self.gauge_canvas.update()
        w = self.gauge_canvas.winfo_width() or 360
        self._draw_rounded_rect(self.gauge_canvas, 0, 0, w, 18, radius=9, fill=DARK["surface"])
        fw = int(pct/100*w)
        if fw > 4:
            self._draw_rounded_rect(self.gauge_canvas, 0, 0, fw, 18, radius=9, fill=DARK["success"])

        for ww in self.stats_inner.winfo_children():
            ww.destroy()
        counts: dict[str,int] = {}
        for d in self.rows:
            s = d.get("Status","?")
            counts[s] = counts.get(s, 0) + 1
        for status in STATUS_ORDER:
            c = counts.get(status, 0)
            if c == 0:
                continue
            bg       = self.status_colors.get(status, DARK["card"])
            icon     = STATUS_ICONS.get(status,"")
            item_pct = c/total*100 if total else 0
            row = tk.Frame(self.stats_inner, bg=DARK["card"])
            row.pack(fill=tk.X, pady=2)
            sw = tk.Canvas(row, width=12, height=12, bg=DARK["card"], highlightthickness=0)
            sw.pack(side=tk.LEFT, padx=(0,4))
            sw.create_rectangle(0,0,12,12,fill=bg,outline="")
            tk.Label(row, text=f"{icon} {status}", font=(self._family,8),
                     bg=DARK["card"], fg=DARK["text"], width=16, anchor=tk.W).pack(side=tk.LEFT)
            bar_w = 80
            bf = tk.Frame(row, bg=DARK["surface"], width=bar_w, height=8)
            bf.pack(side=tk.LEFT, padx=3)
            bf.pack_propagate(False)
            tk.Frame(bf, bg=bg, width=max(2, int(item_pct/100*bar_w))).pack(side=tk.LEFT, fill=tk.Y)
            tk.Label(row, text=f"{c}  ({item_pct:.0f}%)", font=(self._family,7),
                     bg=DARK["card"], fg=DARK["text_dim"]).pack(side=tk.RIGHT)
        tk.Frame(self.stats_inner, bg=DARK["border"], height=1).pack(fill=tk.X, pady=(5,3))
        tk.Label(self.stats_inner, text=f"Total: {total} registry entries",
                 font=(self._family,8,"bold"), bg=DARK["card"], fg=DARK["text"]).pack(anchor=tk.W)

        # Multi-agent folder collision check
        active_items = [d for d in self.rows if d.get("Status") in ("CONSTRUCTION", "NEXT")]
        folder_workers: dict[str, set[str]] = {}
        for item in active_items:
            path = item.get("Path", "")
            owner = item.get("Owner", "") or "Unassigned"
            if "/" in path:
                parent = "/".join(path.split("/")[:-1])
                folder_workers.setdefault(parent, set()).add(owner)
                
        collisions = {folder: owners for folder, owners in folder_workers.items() if len(owners) > 1}
        if collisions:
            tk.Frame(self.stats_inner, bg=DARK["border"], height=1).pack(fill=tk.X, pady=(8,4))
            tk.Label(self.stats_inner, text="⚠️ Folder Collision Alert!",
                     font=(self._family, 8, "bold"), bg=DARK["card"], fg=DARK["danger"]).pack(anchor=tk.W)
            for folder, owners in collisions.items():
                owner_list = ", ".join(owners)
                lbl = tk.Label(self.stats_inner, text=f"• {folder}/\n  → Owners: {owner_list}",
                               font=(self._family, 7), bg=DARK["card"], fg=DARK["text_dim"],
                               justify=tk.LEFT, anchor=tk.W)
                lbl.pack(anchor=tk.W, pady=1)

        self._refresh_summary_tab()

    # ═══════════════════════════════════════════════════════════════════
    # Selection handlers
    # ═══════════════════════════════════════════════════════════════════
    def _on_tree_select(self, event):
        sel = self.tree.selection()
        if not sel:
            self.editor_fields_frame.pack_forget()
            self.editor_placeholder.pack(fill=tk.BOTH, expand=True)
            return
        iid = sel[0]
        if iid in ("__root__",):
            self.editor_fields_frame.pack_forget()
            self.editor_placeholder.pack(fill=tk.BOTH, expand=True)
            return
        if iid.startswith("v_"):
            self.editor_placeholder.pack_forget()
            self.editor_fields_frame.pack(fill=tk.BOTH, expand=True)
            self._load_virtual_folder_panel(iid[2:])
        else:
            self.editor_placeholder.pack_forget()
            self.editor_fields_frame.pack(fill=tk.BOTH, expand=True)
            self._load_detail_panel(int(iid))

    def _load_virtual_folder_panel(self, path):
        self.current_editor_idx = None
        name = path.split("/")[-1]
        self.sel_path_lbl.config(text=name+"/")
        self.sel_icon_lbl.config(text="📁", fg=DARK["accent"])
        self.sel_status_badge.config(text="⚙️ Scaffold", bg=DARK["border"])
        self.sel_meaning_lbl.config(text="Virtual parent folder — not in registry. Add it via ➕ Add Node.")

        for key, (widget, var) in self.editor_inputs.items():
            if isinstance(widget, ttk.Combobox):
                var.set("FROZEN"); widget.configure(state="disabled")
            elif isinstance(widget, tk.Entry):
                var.set(""); widget.configure(state="disabled")
            elif isinstance(widget, tk.Text):
                widget.configure(state="normal")
                widget.delete("1.0", tk.END)
                widget.insert("1.0", f"Parent folder: {path}")
                widget.configure(state="disabled")

        self.apply_btn.configure(state="disabled", text="⚙️ Read Only — Scaffold")
        self.editor_disk_lbl.config(text="📁 Virtual folder node", fg=DARK["text_dim"])
        self.stub_btn.pack_forget()
        self.open_editor_btn.pack_forget()
        self.llm_btn_prompt.configure(state="disabled")
        self.llm_btn_context.configure(state="disabled")
        self.llm_btn_path.configure(state="disabled")
        self._update_preview(None, None)

    def _load_detail_panel(self, idx):
        self.current_editor_idx = idx
        d      = self.rows[idx]
        path   = d.get("Path","")
        etype  = d.get("Type","")
        status = d.get("Status","")

        self.sel_path_lbl.config(text=Path(path).name)
        icon = "📁" if etype.lower() in ("folder","repo") else ("🧪" if etype.lower()=="test" else "📄")
        self.sel_icon_lbl.config(text=icon, fg=DARK["accent"])

        bg   = self.status_colors.get(status, DARK["card"])
        sico = STATUS_ICONS.get(status,"")
        self.sel_status_badge.config(text=f"{sico} {status}", bg=bg)

        # Show the meaning + action clearly
        if status == "FROZEN":
            if path in self.unlocked_frozen_paths:
                self.sel_meaning_lbl.config(
                    text="🔓 SECTION UNLOCKED TEMPORARILY\n→ Edit carefully or mark as active when done."
                )
                self.toggle_edit_btn.configure(
                    text="🔓 Section Unlocked — Adjust Registry Fields...",
                    bg=DARK["surface"],
                    fg=DARK["text_dim"]
                )
            else:
                self.sel_meaning_lbl.config(
                    text="❄️ SECTION FROZEN / LOCKED\n→ Locked to prevent modifications. Click button below to unlock."
                )
                self.toggle_edit_btn.configure(
                    text="🔒 Section Frozen — Click to Unlock...",
                    bg=DARK["surface"],
                    fg=DARK["accent"]
                )
                if hasattr(self, "edit_fields_container") and self.edit_fields_container.winfo_ismapped():
                    self.edit_fields_container.pack_forget()
        else:
            if status in STATUS_WORKFLOW:
                short, meaning, action = STATUS_WORKFLOW[status]
                self.sel_meaning_lbl.config(
                    text=f"{short}  —  {meaning}\n→ What to do: {action}")
            else:
                self.sel_meaning_lbl.config(text="")
            self.toggle_edit_btn.configure(
                text="⚙️ Adjust Registry Fields...",
                bg=DARK["surface"],
                fg=DARK["text_dim"]
            )

        # Populate read-only view
        self.lbl_view_meta.config(text=f"👤 Owner: {d.get('Owner', 'Unassigned')}  •  🎯 Phase: {d.get('Phase', 'N/A')}")
        self.lbl_view_purpose.config(text=d.get("Purpose", "") or "Not specified")
        self.lbl_view_rule.config(text=d.get("Build Rule", "") or "Not specified")
        self.lbl_view_notes.config(text=d.get("Notes", "") or "No notes added")

        self.apply_btn.configure(state="normal", text="💾 Apply Changes")
        self.llm_btn_prompt.configure(state="normal")
        self.llm_btn_context.configure(state="normal")
        self.llm_btn_path.configure(state="normal")

        for key, (widget, var) in self.editor_inputs.items():
            val = d.get(key,"") or ""
            if isinstance(widget, ttk.Combobox):
                widget.configure(state="readonly"); var.set(val)
            elif isinstance(widget, tk.Entry):
                widget.configure(state="normal"); var.set(val)
            elif isinstance(widget, tk.Text):
                widget.configure(state="normal")
                widget.delete("1.0",tk.END); widget.insert("1.0", val)

        project_root = Path(self.linked_project_root) if self.linked_project_root else (Path(self.xlsx_path).parent if self.xlsx_path else None)
        fp = (project_root / path) if project_root else None
        if etype.lower() in ("folder","repo"):
            self.editor_disk_lbl.config(text="📁 Directory node", fg=DARK["text_dim"])
            self.stub_btn.pack_forget(); self.open_editor_btn.pack_forget()
        else:
            if fp and fp.is_file():
                try:
                    sz = fp.stat().st_size
                    sz_str = f'{sz} B' if sz < 1024 else f'{sz/1024:.1f} KB'
                    # Only show arrows on a real measured change vs Git baseline. No stub labels.
                    growth_text = ''
                    orig_sz = self._get_cached_original_size(path)
                    if orig_sz is not None:
                        diff = sz - orig_sz
                        if diff > 50:
                            diff_str = f'+{diff} B' if diff < 1024 else f'+{diff/1024:.1f} KB'
                            growth_text = f'  ▲ {diff_str} since first commit'
                        elif diff < -50:
                            abs_diff = abs(diff)
                            diff_str = f'-{abs_diff} B' if abs_diff < 1024 else f'-{abs_diff/1024:.1f} KB'
                            growth_text = f'  ▼ {diff_str} since first commit'
                    self.editor_disk_lbl.config(
                        text=f'💾 File exists on disk: {sz_str}{growth_text}',
                        fg=DARK['success'])
                except Exception:
                    self.editor_disk_lbl.config(text="💾 File exists on disk", fg=DARK["success"])
                self.stub_btn.pack_forget(); self.open_editor_btn.pack(fill=tk.X, pady=2)
            else:
                self.editor_disk_lbl.config(text="✗  File missing — stub needed", fg=DARK["danger"])
                self.stub_btn.pack(fill=tk.X, pady=2); self.open_editor_btn.pack_forget()

        # Toggle git changes panel dynamically
        git_state = self.git_changed_files.get(path)
        if git_state:
            self.git_frame.pack(fill=tk.X, padx=12, pady=(4, 4))
            self.git_title.config(text=f"🔄 Git {git_state.capitalize()} Detected")
            self.git_desc_lbl.config(
                text=f"Coding round changes found on disk.\nSelect a workflow status to accept:"
            )
        else:
            self.git_frame.pack_forget()

        # Update preview
        self._update_preview(path, etype)

    def _apply_editor_changes(self):
        if self.current_editor_idx is None:
            return
        d  = self.rows[self.current_editor_idx]
        old_status = d.get("Status", "")
        path = d.get("Path", "")
        
        # Enforce frozen lock on save
        if old_status == "FROZEN" and path not in self.unlocked_frozen_paths:
            ans = messagebox.askyesno(
                "❄️ Section Frozen / Locked",
                f"The file '{Path(path).name}' is marked as FROZEN to protect it from accidental changes.\n\nDo you want to unlock it to save these modifications?",
                icon="warning"
            )
            if not ans:
                return
            self.unlocked_frozen_paths.add(path)
            
        ns = ""
        for key, (widget, var) in self.editor_inputs.items():
            if isinstance(widget, ttk.Combobox):
                d[key] = var.get()
                if key == "Status": ns = var.get()
            elif isinstance(widget, tk.Entry):
                d[key] = widget.get().strip()
            elif isinstance(widget, tk.Text):
                d[key] = widget.get("1.0","end-1c").strip()
        if ns:
            meta = STATUS_META.get(ns, {})
            d["Color"] = meta.get("label",""); d["Lock State"] = meta.get("meaning","")
            if ns != old_status:
                self._trigger_status_change_hook(self.current_editor_idx, old_status, ns)
        self.dirty = True
        self.sb_right.config(text="● Unsaved changes", fg=DARK["danger"])
        sel = self.tree.selection()
        self._repopulate(); self._refresh_stats()
        if sel:
            try:
                self.tree.selection_set(sel); self.tree.focus(sel[0])
                self._load_detail_panel(int(sel[0]))
            except Exception:
                pass
        self.apply_btn.config(text="✅ Applied!", bg=DARK["success"], fg="#07101E")
        self.root.after(1500, lambda: self.apply_btn.config(text="💾 Apply Changes",
                                                             bg=DARK["accent"], fg="#07101E"))

    # ─────────────────────────────────────────────────────────────────
    # Keyboard status shortcuts
    # ─────────────────────────────────────────────────────────────────
    def _on_key_press(self, event):
        sel = self.tree.selection()
        if not sel:
            return
        iid = sel[0]
        if iid.startswith("v_") or iid == "__root__":
            return
        km = {"1":"CONSTRUCTION","c":"CONSTRUCTION",
              "2":"NEXT",        "n":"NEXT",
              "3":"SHIP_READY",  "s":"SHIP_READY",
              "4":"ACTIVE_V1",   "a":"ACTIVE_V1",
              "5":"FROZEN",      "f":"FROZEN",
              "6":"HUMAN_GATE",  "h":"HUMAN_GATE",
              "7":"PLACEHOLDER",  "t":"PLACEHOLDER",
              "8":"ROADMAP",     "r":"ROADMAP",
              "9":"LEGACY",      "l":"LEGACY"}
        if event.char.lower() in km:
            self._change_status_immediate(int(iid), km[event.char.lower()])

    def _change_status_immediate(self, idx, new_status):
        d = self.rows[idx]
        if d.get("Status") == new_status:
            return
        old_status = d.get("Status", "")
        path = d.get("Path", "")
        
        # Enforce lock on FROZEN sections
        if old_status == "FROZEN" and path not in self.unlocked_frozen_paths:
            ans = messagebox.askyesno(
                "❄️ Section Frozen / Locked",
                f"The file '{Path(path).name}' is marked as FROZEN to protect it from accidental changes.\n\nDo you want to unlock it to change its status to '{new_status}'?",
                icon="warning"
            )
            if not ans:
                return
            self.unlocked_frozen_paths.add(path)
            
        d["Status"] = new_status
        meta = STATUS_META.get(new_status, {})
        d["Color"] = meta.get("label",""); d["Lock State"] = meta.get("meaning","")
        self.dirty = True
        self.sb_right.config(text="● Unsaved", fg=DARK["danger"])
        
        self._trigger_status_change_hook(idx, old_status, new_status)
        
        sel = self.tree.selection()
        self._repopulate(); self._refresh_stats()
        if sel:
            try:
                self.tree.selection_set(sel); self.tree.focus(sel[0])
                self._load_detail_panel(int(sel[0]))
            except Exception:
                pass

    # ─────────────────────────────────────────────────────────────────
    # Double-click status picker
    # ─────────────────────────────────────────────────────────────────
    def _on_dblclick(self, event):
        iid = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if not iid:
            return
        
        self.tree.selection_set(iid)
        self.tree.focus(iid)
        
        if col in ("#1", "#2"):
            self._status_popup(iid)
        else:
            if not iid.startswith("v_") and iid != "__root__":
                try:
                    idx = int(iid)
                    d = self.rows[idx]
                    if d.get("Type", "").lower() not in ("folder", "repo"):
                        self._open_file_in_editor(d.get("Path", ""))
                except Exception:
                    pass

    def _status_popup(self, iid):
        if iid.startswith("v_") or iid == "__root__":
            return
        idx  = int(iid)
        bbox = self.tree.bbox(iid, column="Status")
        if not bbox:
            return
        x, y, w, h = bbox

        popup = tk.Toplevel(self.root)
        popup.overrideredirect(True)
        popup.configure(bg=DARK["card"], highlightthickness=2,
                        highlightbackground=DARK["accent"])
        abs_x = self.tree.winfo_rootx() + x
        abs_y = self.tree.winfo_rooty() + y + h
        popup.geometry(f"260x{len(STATUS_ORDER)*32+8}+{abs_x}+{abs_y}")

        for status in STATUS_ORDER:
            bg   = self.status_colors.get(status, "#888")
            icon = STATUS_ICONS.get(status,"")
            short, meaning, _ = STATUS_WORKFLOW[status]
            r2,g2,b2 = int(bg[1:3],16), int(bg[3:5],16), int(bg[5:7],16)
            fg = "#0A0A0A" if (0.2126*r2+0.7152*g2+0.0722*b2)>80 else "#1A1A1A"
            b = tk.Button(popup,
                text=f" {icon}  {short:16s}  {meaning[:18]}",
                font=(self._family, 8), bg=bg, fg=fg,
                relief="flat", bd=0, anchor=tk.W, padx=8, pady=3,
                activebackground=DARK["accent_hover"], cursor="hand2",
                command=lambda s=status, p=popup: self._popup_select(idx, s, p))
            b.pack(fill=tk.X, padx=2, pady=1)

        popup.bind("<FocusOut>", lambda e: popup.destroy())
        popup.focus_set(); popup.grab_set()

    def _popup_select(self, idx, new_status, popup):
        popup.destroy()
        self._change_status_immediate(idx, new_status)

    # ─────────────────────────────────────────────────────────────────
    # Right-click context menu
    # ─────────────────────────────────────────────────────────────────
    def _show_context_menu(self, event):
        iid = self.tree.identify_row(event.y)
        if not iid or iid == "__root__":
            return
        self.tree.selection_set(iid); self.tree.focus(iid)
        self.menu.delete(0, tk.END)

        if iid.startswith("v_"):
            path = iid[2:]
            self.menu.add_command(label="🔗 Copy Folder Path",
                                  command=lambda: self._copy_to_clipboard(path))
            self.menu.post(event.x_root, event.y_root)
            return

        idx      = int(iid)
        row_data = self.rows[idx]
        path     = row_data.get("Path","")
        rule     = row_data.get("Build Rule","")
        etype    = row_data.get("Type","").lower()
        fp       = (Path(self.xlsx_path).parent/path) if self.xlsx_path else None

        self.menu.add_command(label="🔗 Copy Path",
                              command=lambda: self._copy_to_clipboard(path))
        self.menu.add_command(label="📋 Copy Build Rule",
                              command=lambda: self._copy_to_clipboard(rule or "No rule defined."))
        self.menu.add_command(label="🤖 Copy AI Prompt",
                              command=lambda: self._copy_prompt_for_idx(idx))
        self.menu.add_command(label="📝 Copy Context Block",
                              command=lambda: self._copy_context_block_for_idx(idx))
        self.menu.add_separator()

        if etype not in ("folder","repo"):
            if fp and fp.exists():
                self.menu.add_command(label="✏️ Open in Editor", command=self._open_selected_in_editor)
            else:
                self.menu.add_command(label="✨ Create Stub on Disk", command=self._create_stub_for_selected)
            self.menu.add_separator()

        sm = tk.Menu(self.menu, tearoff=0, bg=DARK["card"], fg=DARK["text"],
                     activebackground=DARK["accent"], activeforeground="#07101E")
        for s in STATUS_ORDER:
            icon = STATUS_ICONS.get(s,"")
            short = STATUS_WORKFLOW[s][0]
            sm.add_command(label=f"{icon} {short}  — {s}",
                           command=lambda st=s: self._change_status_immediate(idx, st))
        self.menu.add_cascade(label="🎨 Change Status...", menu=sm)
        self.menu.post(event.x_root, event.y_root)

    # ═══════════════════════════════════════════════════════════════════
    # Clipboard / LLM helpers
    # ═══════════════════════════════════════════════════════════════════
    def _copy_to_clipboard(self, text):
        self.root.clipboard_clear(); self.root.clipboard_append(text)
        self.sb_right.config(text="📋 Copied!", fg=DARK["success"])
        self.root.after(2000, lambda: self.sb_right.config(text="", fg=DARK["accent"]))

    def _copy_path_for_selected(self):
        if self.current_editor_idx is not None:
            self._copy_to_clipboard(self.rows[self.current_editor_idx].get("Path",""))

    def _copy_selected_paths(self):
        sel = self.tree.selection()
        if not sel:
            return
        paths = []
        for iid in sel:
            if iid == "__root__":
                continue
            if iid.startswith("v_"):
                paths.append(iid[2:])
            else:
                try:
                    idx = int(iid)
                    paths.append(self.rows[idx].get("Path", ""))
                except Exception:
                    pass
        if paths:
            text = "\n".join(paths)
            self._copy_to_clipboard(text)
            self.sb_right.config(text=f"📋 Copied {len(paths)} path(s)!", fg=DARK["success"])

    def _copy_prompt_for_selected(self):
        if self.current_editor_idx is not None:
            self._copy_prompt_for_idx(self.current_editor_idx)

    def _copy_context_block_for_selected(self):
        if self.current_editor_idx is not None:
            self._copy_context_block_for_idx(self.current_editor_idx)

    def _copy_prompt_for_idx(self, idx):
        d = self.rows[idx]
        prompt = (
            f"Please implement the following {d.get('Type','file')} in our codebase:\n"
            f"**File Path**: `{d.get('Path','')}`\n"
            f"**Purpose**: {d.get('Purpose','') or 'Not specified.'}\n"
            f"**Build Rule**: {d.get('Build Rule','') or 'Follow clean design principles.'}\n"
        )
        if d.get("Notes"):
            prompt += f"**Notes**: {d['Notes']}\n"
        prompt += "\nWrite professional production-ready code. Let's go! 🚀"
        self._copy_to_clipboard(prompt)

    def _copy_context_block_for_idx(self, idx):
        d = self.rows[idx]
        block = (
            f"| Property | Value |\n| --- | --- |\n"
            f"| **Path** | `{d.get('Path','')}` |\n"
            f"| **Type** | `{d.get('Type','')}` |\n"
            f"| **Status** | `{d.get('Status','')}` |\n"
            f"| **Phase** | `{d.get('Phase','N/A')}` |\n"
            f"| **Owner** | `{d.get('Owner','Unassigned')}` |\n"
            f"| **Purpose** | `{d.get('Purpose','N/A')}` |\n"
            f"| **Build Rule** | `{d.get('Build Rule','N/A')}` |\n"
            f"| **Notes** | `{d.get('Notes','N/A')}` |\n"
        )
        self._copy_to_clipboard(block)

    def _export_llm_roadmap(self):
        if not self.rows:
            messagebox.showinfo("Empty","No roadmap data loaded.")
            return
        name = Path(self.xlsx_path).name if self.xlsx_path else "Roadmap"
        md = [f"# 🗺️ Project Roadmap & Codebase Skeleton",
              f"Parsed from: `{name}`\n"]
        active = [d for d in self.rows if d.get("Status") in ("CONSTRUCTION","NEXT")]
        if active:
            md.append("## ⚡ Active Sprint Work Queue")
            for item in active:
                s = item.get("Status"); p = item.get("Path")
                md.append(f"- `[ ]` `{p}` **[{s}]**")
                if item.get("Purpose"):
                    md.append(f"  - *Purpose*: {item['Purpose']}")
            md.append("")
        md.append("## 📁 Registry Outline")
        for d in sorted(self.rows, key=lambda x: x.get("Path","")):
            path  = d.get("Path",""); status = d.get("Status","ROADMAP")
            etype = d.get("Type","file"); purpose = d.get("Purpose","")
            depth = path.count("/"); icon = "📁" if etype.lower()=="folder" else "📄"
            name2 = path.split("/")[-1]
            line  = f"{'  '*depth}- {icon} `{name2}` **[{status}]**"
            if purpose: line += f" — *{purpose}*"
            md.append(line)
        md.append("\n## 🎨 Palette Reference")
        for s in STATUS_ORDER:
            meta = STATUS_META[s]
            md.append(f"- **{s}**: {meta['meaning']} — *{meta['action']}*")
        self._show_text_dialog("📋 LLM Roadmap Export", "\n".join(md))

    def _show_text_dialog(self, title, content):
        dlg = tk.Toplevel(self.root)
        dlg.title(title); dlg.geometry("980x760")
        dlg.configure(bg=DARK["bg"]); dlg.transient(self.root); dlg.grab_set()
        tk.Label(dlg, text=title, font=(self._family,14,"bold"),
                 bg=DARK["bg"], fg=DARK["accent"]).pack(anchor=tk.W, padx=20, pady=(15,5))
        tf = tk.Frame(dlg, bg=DARK["surface"], highlightthickness=1,
                      highlightbackground=DARK["border"])
        tf.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)
        txt = tk.Text(tf, font=self.fm, bg=DARK["surface"], fg=DARK["text"],
                      insertbackground=DARK["text"], relief="flat", bd=0, wrap=tk.WORD)
        vsb = ttk.Scrollbar(tf, orient=tk.VERTICAL, command=txt.yview,
                            style="Dk.Vertical.TScrollbar")
        txt.configure(yscrollcommand=vsb.set)
        txt.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        txt.insert("1.0", content); txt.configure(state=tk.DISABLED)
        bf = tk.Frame(dlg, bg=DARK["bg"])
        bf.pack(fill=tk.X, padx=20, pady=15)
        def copy_all():
            self._copy_to_clipboard(content)
            cb.configure(text="✅ Copied!"); dlg.after(2000, lambda: cb.configure(text="📋 Copy All"))
        cb = tk.Button(bf, text="📋 Copy All", font=(self._family,10,"bold"),
                       bg=DARK["accent"], fg="#07101E", relief="flat", bd=0,
                       padx=20, pady=8, cursor="hand2", command=copy_all)
        cb.pack(side=tk.LEFT)
        tk.Button(bf, text="Close", font=(self._family,10), bg=DARK["card"], fg=DARK["text"],
                  relief="flat", bd=1, padx=20, pady=8, cursor="hand2",
                  command=dlg.destroy).pack(side=tk.RIGHT)

    # ─────────────────────────────────────────────────────────────────
    # Zoom / filter / search
    # ─────────────────────────────────────────────────────────────────
    def _zoom(self, delta):
        nv = self.zoom_level + delta
        if -5 <= nv <= 10:
            self.zoom_level = nv; self._refresh_fonts()

    def _zoom_reset(self):
        self.zoom_level = 0; self._refresh_fonts()

    def _set_filter(self, status):
        if status == "ALL" or self.active_filter == status:
            self.active_filter = None
        else:
            self.active_filter = status
        for s, btn in self.pills.items():
            active = (s == "ALL" and self.active_filter is None) or (s == self.active_filter)
            btn.configure(relief="sunken" if active else "flat",
                          bd=2 if active else 0)
        self._repopulate()

    def _focus_search(self):
        self.search_var.set(""); self.search_entry.focus_set()

    # ─────────────────────────────────────────────────────────────────
    # Close
    # ─────────────────────────────────────────────────────────────────
    def _on_close(self):
        if self.dirty:
            ans = messagebox.askyesnocancel("Unsaved Changes",
                "You have unsaved changes.\n\nSave before closing?")
            if ans is None: return
            if ans: self._save()
        self.root.destroy()

    # ─────────────────────────────────────────────────────────────────
    # Git Integration Helpers & Actions
    # ─────────────────────────────────────────────────────────────────
    def _refresh_git_status(self):
        """Scans the linked project root for modified/untracked files via git status."""
        self.git_changed_files = {}
        self.git_head_sizes = {}
        if not self.linked_project_root or not os.path.isdir(self.linked_project_root):
            return
        
        git_dir = Path(self.linked_project_root) / ".git"
        if not git_dir.exists():
            return
            
        import subprocess
        try:
            # Run git status --porcelain in the linked project root
            res = subprocess.run(
                ["git", "status", "--porcelain"],
                cwd=self.linked_project_root,
                capture_output=True,
                text=True,
                check=True
            )
            for line in res.stdout.splitlines():
                if len(line) < 4:
                    continue
                status_code = line[:2].strip()
                rel_path = line[3:].strip().replace("\\", "/")
                # Normalize path (remove quotes if any)
                if rel_path.startswith('"') and rel_path.endswith('"'):
                    rel_path = rel_path[1:-1]
                
                if "A" in status_code or "?" in status_code:
                    self.git_changed_files[rel_path] = "untracked"
                elif "M" in status_code:
                    self.git_changed_files[rel_path] = "modified"
                    
            # For modified files, query their HEAD baseline size
            for path, state in self.git_changed_files.items():
                if state == "modified":
                    try:
                        res_sz = subprocess.run(
                            ["git", "cat-file", "-s", f"HEAD:{path}"],
                            cwd=self.linked_project_root,
                            capture_output=True,
                            text=True
                        )
                        if res_sz.returncode == 0:
                            self.git_head_sizes[path] = int(res_sz.stdout.strip())
                    except Exception:
                        pass
        except Exception as e:
            print(f"Git status check failed: {e}")

        # Sync current Git branch & unmerged main landing diffs
        self._refresh_branch_status()

    def _periodic_git_refresh(self):
        try:
            self._refresh_git_status()
            self._repopulate()
            self._update_git_sync_card()
            self._refresh_summary_tab()
            if self.current_editor_idx is not None:
                self._load_detail_panel(self.current_editor_idx)
        except Exception:
            pass
        self.root.after(30000, self._periodic_git_refresh)

    def _on_link_project_click(self):
        if not self.xlsx_path:
            messagebox.showinfo("No Workbook", "Please open an Excel workbook first.")
            return
        initial = self.linked_project_root or os.path.dirname(self.xlsx_path)
        folder = filedialog.askdirectory(
            title="Link actual Project Root Folder (where code lives)",
            initialdir=initial,
            parent=self.root
        )
        if folder:
            self.linked_project_root = os.path.normpath(folder)
            # Save to config
            project_roots = self._config.setdefault("project_roots", {})
            abs_xlsx = str(Path(self.xlsx_path).resolve())
            project_roots[abs_xlsx] = self.linked_project_root
            self._persist_config()
            
            # Refresh
            self._refresh_git_status()
            self._repopulate()
            self._refresh_stats()
            self._update_link_label()
            self._update_git_sync_card()
            self.sb_left.config(text=f"Linked Project Root: {self.linked_project_root}")

    def _update_link_label(self):
        if not self.xlsx_path:
            if hasattr(self, "link_lbl"):
                self.link_lbl.pack_forget()
            return
        if hasattr(self, "link_lbl"):
            self.link_lbl.pack(side=tk.LEFT, padx=(15, 0))
            if self.linked_project_root:
                name = Path(self.linked_project_root).name
                branch_info = f" [{self.current_git_branch}]" if self.current_git_branch else ""
                self.link_lbl.config(text=f"🔗 Linked: {name}{branch_info}")
            else:
                self.link_lbl.config(text="🔗 Link Project Root")

    def _manual_git_refresh(self):
        self._refresh_git_status()
        self._repopulate()
        self._refresh_stats()
        self._update_git_sync_card()
        if self.current_editor_idx is not None:
            self._load_detail_panel(self.current_editor_idx)
        self.sb_left.config(text="Git status refreshed manually.")

    def _update_git_sync_card(self):
        if not hasattr(self, "git_status_lbl"):
            return
        if not self.linked_project_root:
            self.git_status_lbl.config(text="No active project root linked.", fg=DARK["text_dim"])
            self.auto_sync_btn.configure(state="disabled")
            return
        
        git_dir = Path(self.linked_project_root) / ".git"
        if not git_dir.exists():
            self.git_status_lbl.config(text=f"Linked folder has no .git repository.", fg=DARK["danger"])
            self.auto_sync_btn.configure(state="disabled")
            return
            
        n = len(self.git_changed_files)
        if n == 0:
            self.git_status_lbl.config(text="All registered files are synced with Git (clean).", fg=DARK["success"])
            self.auto_sync_btn.configure(state="disabled")
        else:
            self.git_status_lbl.config(text=f"Found {n} files with Git changes (modified/new).", fg=DARK["warn"])
            self.auto_sync_btn.configure(state="normal")

    def _bulk_promote_git_changes(self):
        if not self.xlsx_path or not self.linked_project_root:
            return
            
        promoted_count = 0
        for i, d in enumerate(self.rows):
            path = d.get("Path", "")
            status = d.get("Status", "")
            if path in self.git_changed_files:
                if status in ("CONSTRUCTION", "NEXT", "PLACEHOLDER"):
                    old_status = status
                    d["Status"] = "SHIP_READY"
                    meta = STATUS_META["SHIP_READY"]
                    d["Color"] = meta["label"]
                    d["Lock State"] = meta["meaning"]
                    promoted_count += 1
                    self._trigger_status_change_hook(i, old_status, "SHIP_READY")
                    
        if promoted_count > 0:
            self.dirty = True
            self._save()
            self._refresh_git_status()
            self._repopulate()
            self._refresh_stats()
            self._update_git_sync_card()
            messagebox.showinfo("Auto-Promote Succeeded", 
                                f"Successfully promoted {promoted_count} modified files to ✅ SHIP_READY (DONE) and saved sheet!")
        else:
            messagebox.showinfo("No Work-in-Progress Changes", 
                                "No registered files in work-in-progress states (BUILD/QUEUE/STUB) were modified in Git.")

    def _mark_selected_ship_ready(self):
        if self.current_editor_idx is not None:
            self._change_status_immediate(self.current_editor_idx, "SHIP_READY")
            self._save()
            self._refresh_git_status()
            self._repopulate()

    def _mark_selected_active_v1(self):
        if self.current_editor_idx is not None:
            self._change_status_immediate(self.current_editor_idx, "ACTIVE_V1")
            self._save()
            self._refresh_git_status()
            self._repopulate()

    def _refresh_branch_status(self):
        """Checks the current branch name and any changes relative to main."""
        self.current_git_branch = None
        self.branch_diff_files = set()
        
        if not self.linked_project_root or not os.path.isdir(self.linked_project_root):
            return
            
        git_dir = Path(self.linked_project_root) / ".git"
        if not git_dir.exists():
            return
            
        import subprocess
        try:
            # Get current branch name
            res = subprocess.run(
                ["git", "rev-parse", "--abbrev-ref", "HEAD"],
                cwd=self.linked_project_root,
                capture_output=True,
                text=True,
                check=True
            )
            self.current_git_branch = res.stdout.strip()
            
            # If branch is not main/master, get diff against main/master
            if self.current_git_branch not in ("main", "master"):
                ref_branch = "main"
                # Quick check if main exists
                check_main = subprocess.run(
                    ["git", "show-ref", "--verify", "refs/heads/main"],
                    cwd=self.linked_project_root,
                    capture_output=True
                )
                if check_main.returncode != 0:
                    check_master = subprocess.run(
                        ["git", "show-ref", "--verify", "refs/heads/master"],
                        cwd=self.linked_project_root,
                        capture_output=True
                    )
                    if check_master.returncode == 0:
                        ref_branch = "master"
                
                diff_res = subprocess.run(
                    ["git", "diff", ref_branch, "--name-status"],
                    cwd=self.linked_project_root,
                    capture_output=True,
                    text=True
                )
                if diff_res.returncode == 0:
                    for line in diff_res.stdout.splitlines():
                        if len(line) < 2:
                            continue
                        parts = line.split(None, 1)
                        if len(parts) == 2:
                            rel_path = parts[1].strip().replace("\\", "/")
                            self.branch_diff_files.add(rel_path)
        except Exception as e:
            print(f"Branch status check failed: {e}")

    def _is_file_a_stub(self, abs_path, rel_path) -> bool:
        """Determines if a file is just a starting skeleton stub or placeholder."""
        try:
            if not abs_path.is_file():
                return True
            sz = abs_path.stat().st_size
            if sz == 0:
                return True
                
            # 1. Check Git growth: if it has grown from baseline committed size by > 100 bytes, it's not a stub
            if rel_path in self.git_head_sizes:
                head_sz = self.git_head_sizes[rel_path]
                if sz > head_sz + 100:
                    return False
                    
            # 2. Check configurable threshold from config
            threshold = self._config.get("stub_threshold_bytes", 400)
            if sz > threshold * 3:
                return False
                
            # 3. Heuristic Comment Ratio & Placeholder Analyzer
            with open(abs_path, "r", encoding="utf-8", errors="replace") as f:
                lines = []
                for _ in range(100):
                    line = f.readline()
                    if not line:
                        break
                    lines.append(line.strip())
            
            non_empty_lines = [l for l in lines if l]
            if not non_empty_lines:
                return True
                
            comment_indicators = ("#", "//", "/*", "*", "<!--", "-->", "---", "═══")
            comment_count = sum(1 for l in non_empty_lines if any(l.startswith(ind) for ind in comment_indicators))
            
            if comment_count / len(non_empty_lines) >= 0.75:
                return True
                
            placeholder_words = ("placeholder", "todo", "follow standard practices", "implement here", "not specified")
            content_lower = "\n".join(non_empty_lines).lower()
            if sz < threshold and any(word in content_lower for word in placeholder_words):
                return True
                
            return sz < threshold
        except Exception:
            return True

    def _get_git_original_size(self, path) -> int | None:
        """Gets the size of the file in its very first Git commit (its original stub/scaffold state)."""
        if not self.linked_project_root or not os.path.isdir(self.linked_project_root):
            return None
            
        git_dir = Path(self.linked_project_root) / ".git"
        if not git_dir.exists():
            return None
            
        import subprocess
        try:
            res_log = subprocess.run(
                ["git", "log", "--format=%H", "--reverse", "--", path],
                cwd=self.linked_project_root,
                capture_output=True,
                text=True
            )
            if res_log.returncode == 0:
                commits = res_log.stdout.strip().splitlines()
                if commits:
                    first_commit = commits[0]
                    res_sz = subprocess.run(
                        ["git", "cat-file", "-s", f"{first_commit}:{path}"],
                        cwd=self.linked_project_root,
                        capture_output=True,
                        text=True
                    )
                    if res_sz.returncode == 0:
                        return int(res_sz.stdout.strip())
        except Exception:
            pass
        return None

    def _get_cached_original_size(self, path) -> int | None:
        if not hasattr(self, "git_original_sizes"):
            self.git_original_sizes = {}
        if path not in self.git_original_sizes:
            sz = self._get_git_original_size(path)
            if sz is not None:
                self.git_original_sizes[path] = sz
            else:
                self.git_original_sizes[path] = None
        return self.git_original_sizes[path]

    def _bulk_promote_deliverables(self):
        if not self.xlsx_path or not self.linked_project_root:
            return
            
        proj_root = Path(self.linked_project_root)
        promoted_count = 0
        for i, d in enumerate(self.rows):
            path = d.get("Path", "")
            status = d.get("Status", "")
            if status in ("CONSTRUCTION", "NEXT", "PLACEHOLDER") and d.get("Type", "").lower() not in ("folder", "repo"):
                abs_path = proj_root / path
                if abs_path.is_file() and abs_path.stat().st_size > 400:
                    old_status = status
                    d["Status"] = "SHIP_READY"
                    meta = STATUS_META["SHIP_READY"]
                    d["Color"] = meta["label"]
                    d["Lock State"] = meta["meaning"]
                    promoted_count += 1
                    self._trigger_status_change_hook(i, old_status, "SHIP_READY")
                    
        if promoted_count > 0:
            self.dirty = True
            self._save()
            self._refresh_git_status()
            self._repopulate()
            self._refresh_stats()
            self._update_git_sync_card()
            messagebox.showinfo("Deliverables Accepted", 
                                f"Successfully accepted {promoted_count} completed files from your agent, promoted them to ✅ SHIP_READY, and saved workbook!")


    def _update_preview(self, path, etype):
        self.preview_txt.configure(state="normal")
        self.preview_txt.delete("1.0", tk.END)
        
        if not path or not self.xlsx_path or etype.lower() in ("folder", "repo"):
            self.preview_type_badge.config(text="")
            self.preview_txt.insert("1.0", "Select a file to inspect its content or changes.")
            self.preview_txt.configure(state="disabled")
            return
            
        project_root = Path(self.linked_project_root) if self.linked_project_root else Path(self.xlsx_path).parent
        fp = project_root / path
        
        # Check if modified in Git
        git_state = self.git_changed_files.get(path)
        
        if git_state == "modified":
            # Show Git Diff!
            self.preview_type_badge.config(text="GIT DIFF", bg=DARK["warn"], fg="#0A0A0A")
            try:
                import subprocess
                res = subprocess.run(
                    ["git", "diff", path],
                    cwd=self.linked_project_root,
                    capture_output=True,
                    text=True
                )
                diff_output = res.stdout.strip()
                if not diff_output:
                    # Maybe it's staged? Try git diff --cached
                    res = subprocess.run(
                        ["git", "diff", "--cached", path],
                        cwd=self.linked_project_root,
                        capture_output=True,
                        text=True
                    )
                    diff_output = res.stdout.strip()
                    
                if diff_output:
                    for line in diff_output.splitlines():
                        if line.startswith("+") and not line.startswith("+++"):
                            self.preview_txt.insert(tk.END, line + "\n", "diff_add")
                        elif line.startswith("-") and not line.startswith("---"):
                            self.preview_txt.insert(tk.END, line + "\n", "diff_del")
                        elif line.startswith("@@"):
                            self.preview_txt.insert(tk.END, line + "\n", "diff_hdr")
                        else:
                            self.preview_txt.insert(tk.END, line + "\n")
                else:
                    self.preview_txt.insert(tk.END, "No unstaged diff. Reading current file content:\n\n")
                    self._read_raw_file(fp)
            except Exception as e:
                self.preview_txt.insert(tk.END, f"Error getting git diff: {e}\n\nReading current file content:\n\n")
                self._read_raw_file(fp)
                
        elif fp.is_file():
            # Show file content
            sz = fp.stat().st_size
            if sz == 0:
                self.preview_type_badge.config(text="EMPTY FILE", bg=DARK["danger"], fg="#E8F0FE")
                self.preview_txt.insert("1.0", "⚠️ File is currently empty (0 bytes)!", "info")
            else:
                self.preview_type_badge.config(text="FILE CODE", bg=DARK["success"], fg="#0A0A0A")
                self._read_raw_file(fp)
        else:
            self.preview_type_badge.config(text="NOT ON DISK", bg=DARK["danger"], fg="#E8F0FE")
            self.preview_txt.insert("1.0", "✗ File does not exist on disk yet.\nUse 'Create Stub on Disk' above to initialize it.", "info")
            
        self.preview_txt.configure(state="disabled")

    def _read_raw_file(self, fp):
        try:
            with open(fp, "r", encoding="utf-8", errors="replace") as f:
                lines = []
                for _ in range(200):
                    line = f.readline()
                    if not line:
                        break
                    lines.append(line)
                
                content = "".join(lines)
                self.preview_txt.insert(tk.END, content)
                
                if f.readline():
                    self.preview_txt.insert(tk.END, "\n\n... (showing first 200 lines, open file in editor to view full code) ...", "info")
        except Exception as e:
            self.preview_txt.insert(tk.END, f"Error reading file: {e}")


# ═══════════════════════════════════════════════════════════════════════
# Entry Point
# ═══════════════════════════════════════════════════════════════════
def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else None
    root = tk.Tk()
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass
    FileMapperApp(root, xlsx)
    root.mainloop()

if __name__ == "__main__":
    main()
